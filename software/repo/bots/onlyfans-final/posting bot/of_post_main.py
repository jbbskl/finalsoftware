# of_post_main.py
# OnlyFans Posting Bot — async Playwright
# Fresh browser per post. Loads cookies. Navigates to OF, clicks "New post",
# and executes a configurable sequence: caption | quiz | poll | media | price | expiration | post_now
# Captions support Excel MOVE workflow (Excel -> archive JSON/Excel -> replenish source when empty).
# All actionable selectors are tagged with: # [NAV] # [STEP] # [BTN] # [FIELD]

import asyncio
import json
import random
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import yaml
from playwright.async_api import async_playwright, Page, Browser, BrowserContext
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

# ========================= 0. UTILS / CONFIG =========================

def log(msg: str) -> None:
    print(f"[OF POST] {msg}", flush=True)

def load_cfg(path: str = "of_config.yaml") -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

async def safe_click(locator, retries: int = 3, delay: float = 0.8):
    """Click with human-like delays and retry logic"""
    last = None
    for _ in range(retries + 1):
        try:
            await locator.wait_for(state="visible", timeout=15000)
            
            # Human-like pre-click delay
            await asyncio.sleep(random.uniform(0.5, 1.2))
            
            await locator.click()
            
            # Human-like post-click delay
            await asyncio.sleep(random.uniform(0.8, 1.5))
            
            return
        except Exception as e:
            last = e
            await asyncio.sleep(delay)
    raise last

def pick_delay_seconds(block: Dict[str, Any]) -> int:
    mode = (block.get("mode") or "none").lower()
    if mode == "fixed":
        return int(block.get("fixed_seconds", 0))
    if mode == "random":
        mn = int(block.get("random_min_seconds", 0))
        mx = int(block.get("random_max_seconds", mn))
        if mx < mn:
            mx = mn
        return random.randint(mn, mx)
    return 0

# ========================= 1. EXCEL MOVE HELPERS =========================

def _ensure_wb_ws(path: str, sheet: str):
    p = Path(path)
    if p.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        ws0 = wb.active
        ws0.title = sheet
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
    return wb, ws

def _read_column_values(path: str, sheet: str, col_letter: str) -> List[str]:
    if not Path(path).exists():
        return []
    wb = load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    col_idx = column_index_from_string(col_letter)
    vals: List[str] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None and str(v).strip() != "":
            vals.append(str(v))
    return vals

def _write_column_values(ws, col_letter: str, values: List[str]):
    col_idx = column_index_from_string(col_letter)
    # clear generously
    for r in range(1, max(ws.max_row, len(values)) + 50):
        ws.cell(row=r, column=col_idx).value = None
    for i, text in enumerate(values, start=1):
        ws.cell(row=i, column=col_idx).value = text

def _pop_from_excel(excel_cfg: Dict[str, Any]) -> str:
    path = excel_cfg["path"]
    sheet = excel_cfg.get("sheet", "Sheet1")
    col = excel_cfg.get("column", "A")
    strategy = (excel_cfg.get("pick_strategy") or "sequential").lower()

    wb, ws = _ensure_wb_ws(path, sheet)
    values = _read_column_values(path, sheet, col)
    if not values:
        raise RuntimeError("Caption Excel is empty.")

    idx = random.randrange(0, len(values)) if strategy == "random" else 0
    chosen = values[idx]
    remaining = [v for i, v in enumerate(values) if i != idx]
    _write_column_values(ws, col, remaining)
    wb.save(path)
    return chosen

def _archive_caption(text: str, archive_cfg: Dict[str, Any], model: str, post_name: str):
    typ = (archive_cfg.get("type") or "json").lower()
    if typ == "excel":
        ex = archive_cfg.get("excel", {})
        apath = ex["path"]; asheet = ex.get("sheet", "Used"); acol = ex.get("column", "A")
        awb, aws = _ensure_wb_ws(apath, asheet)
        existing = _read_column_values(apath, asheet, acol)
        existing.append(text)
        _write_column_values(aws, acol, existing)
        awb.save(apath)
        return

    # JSON archive
    jpath = archive_cfg.get("json_path", "./of_archive/captions_used.json")
    p = Path(jpath)
    data = {"captions": []}
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8")) or {"captions": []}
        except Exception:
            data = {"captions": []}
    data["captions"].append({
        "text": text,
        "model": model,
        "post": post_name,
        "ts": int(time.time())
    })
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def _replenish_if_empty(excel_cfg: Dict[str, Any], archive_cfg: Dict[str, Any]) -> bool:
    path = excel_cfg["path"]; sheet = excel_cfg.get("sheet", "Sheet1"); col = excel_cfg.get("column", "A")
    current = _read_column_values(path, sheet, col)
    if current:
        return False
    if not (archive_cfg.get("replenish_when_empty", True)):
        return False

    typ = (archive_cfg.get("type") or "json").lower()
    refill: List[str] = []
    if typ == "excel":
        ex = archive_cfg.get("excel", {})
        apath = ex["path"]; asheet = ex.get("sheet", "Used"); acol = ex.get("column", "A")
        refill = _read_column_values(apath, asheet, acol)
    else:
        jpath = archive_cfg.get("json_path", "./of_archive/captions_used.json")
        p = Path(jpath)
        if p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8")) or {"captions": []}
                refill = [c.get("text", "") for c in data.get("captions", []) if c.get("text")]
                if archive_cfg.get("after_replenish_clear_archive", False):
                    p.write_text(json.dumps({"captions": []}, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                refill = []

    if not refill:
        return False

    wb, ws = _ensure_wb_ws(path, sheet)
    _write_column_values(ws, col, refill)
    wb.save(path)
    log(f"♻️ Replenished caption Excel from archive ({len(refill)} items).")
    return True

def resolve_caption_cfg(step: Dict[str, Any], post: Dict[str, Any], model: Dict[str, Any], cfg: Dict[str, Any]) -> Dict[str, Any]:
    # Priority: step.excel -> post.caption.excel -> model.caption.excel -> global.caption.excel
    return (step.get("excel")
            or (post.get("caption") or {}).get("excel")
            or (model.get("caption") or {}).get("excel")
            or (cfg.get("caption") or {}).get("excel")
            or {})

def resolve_archive_cfg(step: Dict[str, Any], post: Dict[str, Any], model: Dict[str, Any], cfg: Dict[str, Any]) -> Dict[str, Any]:
    # Priority: step.caption_archive -> post.caption_archive -> model.caption_archive -> global.caption_archive
    return (step.get("caption_archive")
            or post.get("caption_archive")
            or model.get("caption_archive")
            or cfg.get("caption_archive")
            or {})

# Stores context for archiving after a successful Post
class CaptionMoveCtx:
    def __init__(self):
        self.used_text: Optional[str] = None
        self.excel_cfg: Optional[Dict[str, Any]] = None
        self.archive_cfg: Optional[Dict[str, Any]] = None
        self.was_archived: bool = False

# ========================= 2. BROWSER / COOKIES =========================

async def load_session_state(context: BrowserContext, session_path: str):
    """Load full session state (cookies + localStorage + sessionStorage)"""
    p = Path(session_path)
    if not p.exists():
        log(f"No session state at {session_path}; proceeding unauthenticated")
        return
    try:
        # Try to load as full session state first
        session_data = json.loads(p.read_text(encoding="utf-8"))
        if "cookies" in session_data and "origins" in session_data:
            # This is a full session state file
            await context.add_cookies(session_data["cookies"])
            log(f"Loaded session state from {session_path} ({len(session_data['cookies'])} cookies)")
        else:
            # This is just cookies array
            await context.add_cookies(session_data)
            log(f"Loaded cookies from {session_path} ({len(session_data)} cookies)")
    except Exception as e:
        log(f"Failed to load session: {e}")

# Login functions removed - now using manual login approach

async def open_home(cfg: Dict[str, Any], model: Dict[str, Any] = None):
    pw = await async_playwright().start()
    
    # Use persistent browser data directory for session persistence
    browser_data_dir = Path("./browser_data")
    browser_data_dir.mkdir(exist_ok=True)
    
    # Launch browser with persistent user data
    browser: Browser = await pw.chromium.launch_persistent_context(
        user_data_dir=str(browser_data_dir),
        headless=cfg.get("headless", True),
        args=["--window-size=1200,800"],
        viewport={"width": 1200, "height": 800}
    )
    
    # Get the existing page or create new one
    if browser.pages:
        page = browser.pages[0]
    else:
        page = await browser.new_page()

    log('# [NAV] Go to OnlyFans (not homepage to avoid redirects)')
    await page.goto("https://onlyfans.com", wait_until="domcontentloaded", timeout=cfg["timeouts"]["long_ms"])
    
    # Wait for page to load
    await asyncio.sleep(3)
    
    # Check if we're logged in
    login_form = page.locator('form[action*="login"], input[name="email"], input[type="email"]')
    try:
        await login_form.wait_for(state="visible", timeout=5000)
        log("❌ LOGIN REQUIRED: Browser session not authenticated")
        log("🔍 Current URL: " + page.url)
        log("📄 Page title: " + await page.title())
        log("")
        log("🔧 SOLUTION: Run 'python setup_session.py' first to authenticate")
        log("   This will open a persistent browser session for login")
        log("   The session will be saved and reused automatically")
        
        raise RuntimeError("Not authenticated. Please run setup session first: python setup_session.py")
        
    except Exception as e:
        if "Timeout" in str(e):
            # Good! No login form found, we're logged in
            log("✅ Authenticated successfully via persistent browser session")
            log("🔍 Current URL: " + page.url)
        else:
            raise e

    # best-effort cookie banner accept (optional)
    try:
        accept = page.locator('button:has-text("Accept"), button:has-text("I agree")')
        await accept.wait_for(state="visible", timeout=3000)
        await accept.click()
        await asyncio.sleep(2)
        log('# [BTN] Accepted cookie consent')
    except Exception:
        pass

    return pw, browser, page

async def close_browser(pw, browser: Browser):
    # For persistent context, we just close the browser (data is auto-saved)
    await browser.close()
    await pw.stop()

# ========================= 3. CORE FLOW: OPEN "NEW POST" =========================

async def open_new_post(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] "New post" (+)')
    new_post = page.locator('a.m-create-post[ data-name="PostsCreate" ], a[href="/posts/create"]')
    await safe_click(new_post, retries=cfg["retries"]["clicks"])
    await page.wait_for_url("**/posts/create*", timeout=cfg["timeouts"]["long_ms"])

# ========================= 4. COMPONENTS =========================
# ── Caption ──────────────────────────────────────────────────────

async def comp_caption(page: Page, cfg: Dict[str, Any], text: str):
    log('# [FIELD] Caption contenteditable')
    editor = page.locator('div.tiptap.ProseMirror[contenteditable="true"]')
    await editor.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    
    # Human-like interaction with caption field
    await asyncio.sleep(random.uniform(0.5, 1.0))  # Pause before clicking
    await editor.click()
    await asyncio.sleep(random.uniform(0.3, 0.8))  # Pause after clicking
    
    # Type text character by character with human-like delays
    if text:
        await editor.fill("")  # Clear first
        await asyncio.sleep(random.uniform(0.2, 0.5))
        
        for char in text:
            await editor.type(char)
            # Variable typing speed
            delay = random.uniform(0.05, 0.15)
            if char in " .,!?":  # Pause longer on punctuation
                delay += random.uniform(0.1, 0.3)
            await asyncio.sleep(delay)
    
    # Small pause after typing
    await asyncio.sleep(random.uniform(0.5, 1.0))

# ── Quiz (2–10 options, one correct, duration) ───────────────────

async def comp_quiz(page: Page, cfg: Dict[str, Any], quiz: Dict[str, Any]):
    log('# [BTN] Add quiz')
    btn = page.locator('button[at-attr="add_quiz"]')
    await safe_click(btn, retries=cfg["retries"]["clicks"])

    voting_scope = page.locator('div.b-make-post__voting-wrapper, div.b-make-post__voting__body').first
    await voting_scope.wait_for(state="visible", timeout=cfg["timeouts"]["long_ms"])

    desired = max(2, min(int(quiz.get("num_options", 2)), 10))
    add_more = voting_scope.locator('button.new_vote_add_option')

    async def ensure_count():
        for _ in range(12):
            inputs = await voting_scope.locator('input[at-attr="input"]').all()
            current_count = len(inputs)
            if current_count >= desired:
                return
            try:
                log(f"# [BTN] Adding option {current_count + 1} (need {desired} total)")
                
                # Human-like delay before clicking add button
                await asyncio.sleep(random.uniform(0.5, 1.0))
                await add_more.click()
                
                # Human-like delay after adding option
                await asyncio.sleep(random.uniform(0.8, 1.5))
                
            except Exception as e:
                log(f"⚠️ Could not add more options: {e}")
                break

    await ensure_count()

    options: List[str] = quiz.get("options") or []
    inputs = voting_scope.locator('input[at-attr="input"]')
    count = min(desired, await inputs.count(), len(options))
    for i in range(count):
        inp = inputs.nth(i)
        
        # Human-like interaction with each quiz option
        await asyncio.sleep(random.uniform(0.3, 0.7))  # Pause between options
        await inp.click()
        await asyncio.sleep(random.uniform(0.2, 0.5))  # Pause after click
        
        try:
            await inp.fill("")
        except Exception:
            pass
        
        # Type option with human-like speed
        option_text = options[i]
        for char in option_text:
            await inp.type(char)
            await asyncio.sleep(random.uniform(0.08, 0.18))
        
        await asyncio.sleep(random.uniform(0.2, 0.4))  # Pause after typing

    if "correct_index" in quiz:
        idx = int(quiz["correct_index"])
        log(f'# [BTN] Quiz correct option index = {idx + 1}')
        
        # Human-like delay before selecting correct answer
        await asyncio.sleep(random.uniform(0.5, 1.2))
        
        radios = voting_scope.locator('button.m-quiz-swithcer')
        correct_radio = radios.nth(idx)
        
        # Human-like interaction with correct answer selection
        await asyncio.sleep(random.uniform(0.3, 0.7))
        await correct_radio.click()
        await asyncio.sleep(random.uniform(0.5, 1.0))
        
        log(f'✅ Selected correct answer: option {idx + 1}')

    dur = (quiz.get("duration") or "7d").lower()
    # Convert "X days" format to "Xd" format
    dur = dur.replace(" days", "d").replace(" day", "d")
    
    if dur in ("no_limit", "1d", "3d", "7d", "30d"):
        log(f'# [BTN] Quiz duration -> {dur}')
        
        # Human-like delay before setting duration
        await asyncio.sleep(random.uniform(0.5, 1.0))
        
        open_dur = voting_scope.locator('span.b-post-piece__value').first
        try:
            await open_dur.click()
            await asyncio.sleep(random.uniform(0.5, 1.0))
        except Exception:
            pass
        
        # Use text-based selectors instead of IDs to avoid CSS issues
        duration_text_map = {
            "no_limit": "No limit",
            "1d": "1 day",
            "3d": "3 days", 
            "7d": "7 days",
            "30d": "30 days",
        }
        
        duration_text = duration_text_map.get(dur, "7 days")
        target = page.locator(f'button:has-text("{duration_text}")')
        await safe_click(target, retries=cfg["retries"]["clicks"])
        
        # Click Save button to confirm duration selection
        log('# [BTN] Save quiz duration')
        save_btn = page.locator('button:has-text("Save")')
        await safe_click(save_btn, retries=cfg["retries"]["clicks"])
        
        log(f'✅ Set quiz duration to: {dur}')

# ── Poll (2–10 options, duration) ────────────────────────────────

async def comp_poll(page: Page, cfg: Dict[str, Any], poll: Dict[str, Any]):
    log('# [BTN] Add poll')
    btn = page.locator('button[at-attr="add_poll"]')
    await safe_click(btn, retries=cfg["retries"]["clicks"])

    body = page.locator('div.b-make-post__voting__body').first
    await body.wait_for(state="visible", timeout=cfg["timeouts"]["long_ms"])

    desired = max(2, min(int(poll.get("num_options", 2)), 10))
    add_more = body.locator('button.new_vote_add_option')

    async def ensure_count():
        for _ in range(12):
            inputs = await body.locator('input[at-attr="input"]').all()
            current_count = len(inputs)
            if current_count >= desired:
                return
            try:
                log(f"# [BTN] Adding poll option {current_count + 1} (need {desired} total)")
                
                # Human-like delay before clicking add button
                await asyncio.sleep(random.uniform(0.5, 1.0))
                await add_more.click()
                
                # Human-like delay after adding option
                await asyncio.sleep(random.uniform(0.8, 1.5))
                
            except Exception as e:
                log(f"⚠️ Could not add more poll options: {e}")
                break

    await ensure_count()

    options: List[str] = poll.get("options") or []
    inputs = body.locator('input[at-attr="input"]')
    count = min(desired, await inputs.count(), len(options))
    for i in range(count):
        inp = inputs.nth(i)
        
        # Human-like interaction with each poll option
        await asyncio.sleep(random.uniform(0.3, 0.7))  # Pause between options
        await inp.click()
        await asyncio.sleep(random.uniform(0.2, 0.5))  # Pause after click
        
        try:
            await inp.fill("")
        except Exception:
            pass
        
        # Type option with human-like speed
        option_text = options[i]
        for char in option_text:
            await inp.type(char)
            await asyncio.sleep(random.uniform(0.08, 0.18))
        
        await asyncio.sleep(random.uniform(0.2, 0.4))  # Pause after typing

    dur = (poll.get("duration") or "7d").lower()
    # Convert "X days" format to "Xd" format
    dur = dur.replace(" days", "d").replace(" day", "d")
    
    if dur in ("no_limit", "1d", "3d", "7d", "30d"):
        log(f'# [BTN] Poll duration -> {dur}')
        
        # Longer delay before setting duration (so you can see it)
        await asyncio.sleep(random.uniform(1.5, 2.5))
        
        # Click the poll duration span within the b-post-piece container
        try:
            # Look for the exact poll container structure you provided
            poll_container = page.locator('div.b-post-piece.g-pointer-cursor:has(svg[data-icon-name="icon-poll"])')
            duration_span = poll_container.locator('span.b-post-piece__value.g-text-ellipsis')
            
            await duration_span.wait_for(state="visible", timeout=5000)
            await duration_span.click()
            await asyncio.sleep(random.uniform(1.0, 2.0))  # Longer delay to see dropdown
            log('✅ Clicked poll duration span: "7 days"')
            
        except Exception as e:
            log(f'⚠️ Could not find poll duration span: {e}')
            return
        
        # Use text-based selectors instead of IDs to avoid CSS issues
        duration_text_map = {
            "no_limit": "No limit",
            "1d": "1 day",
            "3d": "3 days",
            "7d": "7 days",
            "30d": "30 days",
        }
        
        duration_text = duration_text_map.get(dur, "7 days")
        target = page.locator(f'button:has-text("{duration_text}")')
        
        # Longer delay before clicking duration option
        await asyncio.sleep(random.uniform(1.0, 1.8))
        await safe_click(target, retries=cfg["retries"]["clicks"])
        
        # Longer delay before clicking Save
        await asyncio.sleep(random.uniform(1.2, 2.0))
        
        # Click Save button to confirm duration selection
        log('# [BTN] Save poll duration')
        save_btn = page.locator('button:has-text("Save")')
        await safe_click(save_btn, retries=cfg["retries"]["clicks"])
        
        # Longer delay after saving to see the result
        await asyncio.sleep(random.uniform(1.5, 2.5))
        
        log(f'✅ Set poll duration to: {dur}')

# ── Media from Vault (search folder, select N, add) ───────────────

async def comp_media(page: Page, cfg: Dict[str, Any], media_cfg: Dict[str, Any]):
    log('# [BTN] Add media from vault')
    btn = page.locator('button[at-attr="add_vault_media"]')
    await safe_click(btn, retries=cfg["retries"]["clicks"])

    # Wait for modal to load
    log('# [MODAL] Waiting for vault modal to load')
    modal_body = page.locator('#ModalMediaVault___BV_modal_body_')
    await modal_body.wait_for(state="visible", timeout=cfg["timeouts"]["long_ms"])

    # Support both folder_filter and folder_name for compatibility
    folder_name = media_cfg.get("folder_name") or media_cfg.get("folder_filter", "")
    
    # Check if "all media" filter is requested
    use_all_media = folder_name.lower() in ["all media", "all", ""]
    
    if folder_name and not use_all_media:
        # Only click search button if we need to search for a specific folder
        log('# [BTN] Vault search')
        search_btn = page.locator('div.b-btns-group button[aria-label="Search"]').first
        await safe_click(search_btn, retries=cfg["retries"]["clicks"])
        
        log(f'# [FIELD] Typing search filter: "{folder_name}"')
        
        # After clicking search button, we can directly type (no need to find input)
        await asyncio.sleep(random.uniform(2.0, 3.0))  # Longer wait for search to activate
        
        # Type the search term directly (search input should be focused)
        await page.keyboard.type(folder_name)
        
        # Human-like delay after typing
        await asyncio.sleep(random.uniform(1.5, 2.5))  # Longer delay
        
        # Press Enter to execute search
        log('# [KEY] Pressing Enter to search')
        await page.keyboard.press("Enter")
        
        # Wait for search results to load
        await asyncio.sleep(random.uniform(3.0, 5.0))  # Much longer wait
        log('✅ Search filter applied')

        # [BTN] Click folder by visible name - with 5 second delays as requested
        folder_name = folder_name.strip()
        log(f'# [BTN] Open folder: "{folder_name}"')
        await asyncio.sleep(5.0)  # 5 second delay before folder click
        folder = modal_body.locator(f'div.b-rows-lists__item__name:has-text("{folder_name}")').first
        await safe_click(folder, retries=cfg["retries"]["clicks"])
        await asyncio.sleep(5.0)  # 5 second delay after folder click
        log('✅ Folder opened, waiting for media to load')
    elif use_all_media:
        log('# [FILTER] Using "all media" - vault already shows all media, no search needed')
        # Wait longer for the vault to load all media (already in main view)
        await asyncio.sleep(random.uniform(3.0, 5.0))

    # Try different potential scroll containers
    # First try the modal body itself (might be the actual scrollable area)
    scroll_box = modal_body
    
    # Also get the media container for counting items  
    media_container = modal_body.locator('div.b-feed-content.b-vault-media.g-negative-sides-gaps')
    
    # Wait for media container to be visible
    await media_container.wait_for(state="visible", timeout=cfg["timeouts"]["long_ms"])

    select_n = int(media_cfg.get("count", 1))
    log(f"# [STEP] Selecting {select_n} item(s)")

    async def scroll_more(times=5):
        """Try multiple scrolling approaches to find what works"""
        log(f'# [STEP] Trying different scrolling methods ({times} times)')
        
        for i in range(times):
            try:
                # Method 1: Keyboard scrolling with bigger jumps
                await media_container.focus()
                for _ in range(30):  # More page downs than before
                    await page.keyboard.press('PageDown')
                    await asyncio.sleep(0.05)
                log(f'📜 Keyboard scroll {i+1}/{times}: 30x PageDown')
                
                await asyncio.sleep(1.0)
                
                # Method 2: ScrollIntoView on elements further down
                try:
                    items = await media_container.locator('div[at-attr="checkbox"].checkbox-item').all()
                    if len(items) > 10:
                        await items[-1].scroll_into_view_if_needed()
                        log(f'📜 ScrollIntoView {i+1}/{times}: scrolled to last visible item')
                except Exception:
                    pass
                
                await asyncio.sleep(1.0)
                
                # Method 3: Bigger JavaScript scrolling
                try:
                    await modal_body.evaluate("""
                        (el) => {
                            const scrollers = [
                                el,
                                el.querySelector('.vue-recycle-scroller'),
                                el.querySelector('.b-feed-content'),
                                document.querySelector('#ModalMediaVault___BV_modal_body_')
                            ].filter(Boolean);
                            
                            scrollers.forEach(scroller => {
                                if (scroller && scroller.scrollTop !== undefined) {
                                    scroller.scrollTop += 10000;  // Much bigger scroll jump
                                    scroller.dispatchEvent(new Event('scroll', { bubbles: true }));
                                }
                            });
                        }
                    """)
                    log(f'📜 JavaScript scroll {i+1}/{times}: multiple containers + 10000px')
                except Exception:
                    pass
                
                # Wait for Vue virtual scroller to render items
                await asyncio.sleep(random.uniform(2.0, 3.0))  # Reasonable wait for rendering
                
            except Exception as e:
                log(f'⚠️ Scroll error on attempt {i+1}: {e}')
                pass

    # First, scan what's already loaded without scrolling
    await asyncio.sleep(random.uniform(2.0, 3.0))  # Longer wait for initial load
    checkboxes = await media_container.locator('div[at-attr="checkbox"].checkbox-item').all()
    available_count = len(checkboxes)
    log(f'📊 Initial scan: {available_count} media items already loaded')
    
    # If using "all media", scroll to load more items
    if use_all_media:
        log('🔄 Scrolling to load all available media items...')
        
        last_count = available_count
        no_change_count = 0
        max_attempts = 15
        
        for attempt in range(max_attempts):
            await scroll_more(1)  # One round of scrolling
            await asyncio.sleep(random.uniform(1.5, 2.5))
            
            # Check progress
            checkboxes = await modal_body.locator('div[at-attr="checkbox"].checkbox-item').all()
            new_count = len(checkboxes)
            log(f'📊 Scroll attempt {attempt + 1}: {new_count} media items loaded')
            
            # Check if we're still loading new items
            if new_count == last_count:
                no_change_count += 1
                if no_change_count >= 5:  # Stop if no new items for 5 attempts
                    log(f'✅ No new items loaded for 5 attempts, stopping at {new_count} items')
                    break
            else:
                no_change_count = 0  # Reset counter if we got new items
            
            last_count = new_count
        
        # Final count after scroll attempts
        available_count = new_count
        log(f'📊 Final media count: {available_count}')
    
    selected = 0
    tried_loops = 0
    max_loops = 3  # Reduced loops for selection phase
    
    while selected < select_n and tried_loops < max_loops:
        tried_loops += 1
        
        # Re-scan available items from media container
        checkboxes = await media_container.locator('div[at-attr="checkbox"].checkbox-item').all()
        available_count = len(checkboxes)
        log(f'📊 Found {available_count} media items available for selection (loop {tried_loops})')
        
        if available_count == 0:
            log('⚠️ No media items found - trying light scroll')
            await scroll_more(2)
            continue
        
        # Check if we have enough items already loaded
        if available_count >= select_n:
            log(f'✅ Sufficient media items loaded ({available_count} >= {select_n})')
        
        # Select from available items
        random.shuffle(checkboxes)
        for cb in checkboxes[:select_n - selected]:  # Only select what we need
            try:
                # Human-like media selection
                await asyncio.sleep(random.uniform(0.4, 0.8))  # Think before selecting
                await cb.click()
                selected += 1
                log(f'✅ Selected media item {selected}/{select_n}')
                await asyncio.sleep(random.uniform(0.3, 0.6))  # Shorter pause after selection
            except Exception as e:
                log(f'⚠️ Failed to select media item: {e}')
                continue
        
        # Only scroll if we still need more items and haven't found enough
        if selected < select_n and available_count < select_n * 2:
            log(f'🔄 Need {select_n - selected} more items - trying light scroll...')
            await scroll_more(2)
        else:
            break  # We have enough or scrolling won't help

    # [BTN] Add selected
    log('# [BTN] Add selected media')
    add_btn = modal_body.locator('div.b-row-selected__controls button:has-text("Add")')
    await safe_click(add_btn, retries=cfg["retries"]["clicks"])

# ── Price toggle (and amount if present) ──────────────────────────

async def comp_price(page: Page, cfg: Dict[str, Any], price_cfg: Dict[str, Any]):
    log('# [BTN] Post price')
    price_btn = page.locator('button[at-attr="price_btn"]')
    await safe_click(price_btn, retries=cfg["retries"]["clicks"])
    
    amount = price_cfg.get("amount")
    if amount is None:
        return
        
    log(f'# [FIELD] Typing price amount: ${amount}')
    
    # After clicking price button, input should be active - just type directly
    await asyncio.sleep(random.uniform(0.8, 1.5))  # Wait for price input to activate
    
    # Type price directly (input should be focused after clicking price button)
    price_str = str(amount)
    await page.keyboard.type(price_str)
    
    # Human-like delay after typing
    await asyncio.sleep(random.uniform(0.8, 1.5))
    
    # Look for Save button if price has one
    try:
        save_btn = page.locator('button:has-text("Save")').first
        await save_btn.wait_for(state="visible", timeout=3000)
        log('# [BTN] Save price')
        await safe_click(save_btn, retries=cfg["retries"]["clicks"])
        log(f'✅ Set price to: ${amount}')
    except Exception:
        # No save button found - price might be set automatically
        log(f'✅ Price set to: ${amount} (no save button needed)')

# ── Expiration period ─────────────────────────────────────────────

async def comp_expiration(page: Page, cfg: Dict[str, Any], val: str):
    log('# [BTN] Expiration date menu')
    exp_btn = page.locator('button.b-make-post__expire-period-btn[aria-label="Expiration period"], button.b-make-post__expire-period-btn')
    await safe_click(exp_btn, retries=cfg["retries"]["clicks"])

    val = (val or "").lower()
    # Convert "X days" format to "Xd" format
    val = val.replace(" days", "d").replace(" day", "d")
    
    # Use text-based selectors instead of IDs to avoid CSS issues
    expiration_text_map = {
        "1d": "1 day",
        "3d": "3 days",
        "7d": "7 days",
        "30d": "30 days",
        "no_limit": "No limit",
    }
    
    expiration_text = expiration_text_map.get(val)
    if not expiration_text:
        log(f'⚠️ Unknown expiration "{val}" — skipping selection')
        return

    log(f'# [BTN] Expiration -> {val}')
    tab = page.locator(f'button:has-text("{expiration_text}")')
    await safe_click(tab, retries=cfg["retries"]["clicks"])
    log(f'✅ Set post expiration to: {val}')

# ── Post Now (archives caption if popped) ─────────────────────────

async def comp_post_now(page: Page, cfg: Dict[str, Any], move_ctx: CaptionMoveCtx, model_name: str, post_name: str):
    log('# [BTN] Post now')
    
    # DRY-RUN kill switch
    if cfg.get("dry_run", False):
        log("🧪 DRY-RUN: skipping final POST click")
        return
    
    post_btn = page.locator('button[at-attr="submit_post"]:has-text("Post")')
    await safe_click(post_btn, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(4)
    # Archive the used caption (if any)
    if move_ctx.used_text and move_ctx.excel_cfg and move_ctx.archive_cfg and not move_ctx.was_archived:
        try:
            _archive_caption(move_ctx.used_text, move_ctx.archive_cfg, model_name, post_name)
            move_ctx.was_archived = True
            log("🗂️ Archived used caption.")
        except Exception as e:
            log(f"⚠️ Failed to archive caption: {e}")

# ========================= 5. SEQUENCE ENGINE =========================

def _resolve_caption_text(step: Dict[str, Any], post: Dict[str, Any], model: Dict[str, Any], cfg: Dict[str, Any], move_ctx: CaptionMoveCtx) -> str:
    # If direct text is provided, use it (no MOVE).
    if step.get("text"):
        return str(step["text"])

    # Try Excel MOVE providers.
    excel_cfg = resolve_caption_cfg(step, post, model, cfg)
    if excel_cfg:
        archive_cfg = resolve_archive_cfg(step, post, model, cfg)
        try:
            if _replenish_if_empty(excel_cfg, archive_cfg):
                log("# [STEP] Replenished caption Excel from archive")
            text = _pop_from_excel(excel_cfg)
            move_ctx.used_text = text
            move_ctx.excel_cfg = excel_cfg
            move_ctx.archive_cfg = archive_cfg
            return text
        except Exception as e:
            log(f"⚠️ Caption Excel pop failed: {e}. Falling back to empty caption.")
            return ""
    return ""

async def run_sequence(page: Page, cfg: Dict[str, Any], post: Dict[str, Any], move_ctx: CaptionMoveCtx, model_name: str, post_name: str):
    """
    Handles both formats:
    1. Original: post.sequence = [{ type: "caption", text: "..." }, { type: "quiz", ... }]
    2. Test: post = { sequence: ["caption", "quiz", "post"], caption: {...}, quiz: {...} }
    """
    # Check if it's the new test format or original format
    sequence_list = post.get("sequence", [])
    if sequence_list and isinstance(sequence_list[0], str):
        # New test format: sequence is list of strings, configs are top-level
        for step_name in sequence_list:
            step_name = step_name.lower().strip()
            if step_name == "caption":
                caption_cfg = post.get("caption", {})
                text = caption_cfg.get("text", "")
                if not text:
                    # Try Excel MOVE if no direct text
                    text = _resolve_caption_text(caption_cfg, post, {}, cfg, move_ctx)
                await comp_caption(page, cfg, text)
            elif step_name == "quiz":
                quiz_cfg = post.get("quiz", {})
                await comp_quiz(page, cfg, quiz_cfg)
            elif step_name == "poll":
                poll_cfg = post.get("poll", {})
                await comp_poll(page, cfg, poll_cfg)
            elif step_name == "media":
                media_cfg = post.get("media", {})
                await comp_media(page, cfg, media_cfg)
            elif step_name == "price":
                price_cfg = post.get("price", {})
                await comp_price(page, cfg, price_cfg)
            elif step_name in ("expiration", "expire"):
                # Check if this is a paid post - expiration not allowed for paid posts only
                price_cfg = post.get("price", {})
                has_price = price_cfg.get("amount") is not None and price_cfg.get("amount") > 0
                
                if has_price:
                    log("⚠️ SKIPPING expiration: Not allowed for paid posts (free posts, polls, quizzes can still use expiration)")
                else:
                    expire_cfg = post.get("expire", {}) or post.get("expiration", {})
                    duration = expire_cfg.get("duration", "7d")
                    # Convert "X days" format to "Xd" format
                    duration = duration.replace(" days", "d").replace(" day", "d")
                    log(f"✅ Setting expiration for free post: {duration}")
                    await comp_expiration(page, cfg, duration)
            elif step_name == "post":
                await comp_post_now(page, cfg, move_ctx, model_name, post_name)
            else:
                log(f'⚠️ Unknown step "{step_name}" — skipping')
    else:
        # Original format: sequence is list of dicts with type field
        for step in sequence_list:
            t = (step.get("type") or "").lower().strip()
            if t == "caption":
                text = _resolve_caption_text(step, post, {}, cfg, move_ctx)
                await comp_caption(page, cfg, text)
            elif t == "quiz":
                await comp_quiz(page, cfg, step)
            elif t == "poll":
                await comp_poll(page, cfg, step)
            elif t == "media":
                await comp_media(page, cfg, step)
            elif t == "price":
                await comp_price(page, cfg, step)
            elif t == "expiration":
                await comp_expiration(page, cfg, step.get("value", "7d"))
            elif t == "post_now":
                await comp_post_now(page, cfg, move_ctx, model_name, post_name)
            else:
                log(f'⚠️ Unknown step type "{t}" — skipping')

# ========================= 6. ORCHESTRATION =========================

async def run_post(cfg: Dict[str, Any], model: Dict[str, Any], post: Dict[str, Any]):
    move_ctx = CaptionMoveCtx()
    pw, browser, page = await open_home(cfg, model)
    try:
        await open_new_post(page, cfg)
        await run_sequence(page, cfg, post, move_ctx, model.get("name","model"), post.get("name","post"))
    except Exception as e:
        # If a caption was popped but not archived, restore it to the top of source Excel
        if move_ctx.used_text and move_ctx.excel_cfg and not move_ctx.was_archived:
            try:
                wb, ws = _ensure_wb_ws(move_ctx.excel_cfg["path"], move_ctx.excel_cfg.get("sheet","Sheet1"))
                col = move_ctx.excel_cfg.get("column","A")
                current = _read_column_values(move_ctx.excel_cfg["path"], move_ctx.excel_cfg.get("sheet","Sheet1"), col)
                current.insert(0, move_ctx.used_text)
                _write_column_values(ws, col, current)
                wb.save(move_ctx.excel_cfg["path"])
                log("↩️ Restored popped caption back to Excel due to failure.")
            except Exception as ee:
                log(f"⚠️ Failed to restore caption: {ee}")
        raise e
    finally:
        await close_browser(pw, browser)

async def main():
    cfg = load_cfg()
    models: List[Dict[str, Any]] = cfg.get("models", [])
    pace_all = cfg.get("pace", {}).get("between_posts", {}) or {"mode": "none"}

    for m_i, model in enumerate(models, start=1):
        posts = model.get("posts", [])
        log("=" * 70)
        log(f"MODEL {m_i}/{len(models)}: {model.get('name','(unnamed)')} | posts: {len(posts)}")
        log("=" * 70)
        pace = model.get("pace", {}).get("between_posts", pace_all)

        for p_i, post in enumerate(posts, start=1):
            log(f"--- Post {p_i}/{len(posts)} :: {post.get('name','post')} ---")
            await run_post(cfg, model, post)
            if p_i < len(posts):
                delay = pick_delay_seconds(pace)
                if delay > 0:
                    log(f"⏳ Waiting {delay}s before next post…")
                    await asyncio.sleep(delay)

if __name__ == "__main__":
    asyncio.run(main())
