# main.py
# F2F Posting Bot — async Playwright (multi-post per model + pacing + caption MOVE + per-post fresh browser)
# Each action is tagged: # [NAV], # [STEP], # [BTN], # [FIELD]

import asyncio
import json
import random
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import yaml
from zoneinfo import ZoneInfo
from playwright.async_api import async_playwright, Page, Browser, BrowserContext, TimeoutError as PWTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


# ---------------------------- Logging & Config ----------------------------

def log(msg: str) -> None:
    print(f"[F2F] {msg}", flush=True)

def load_cfg(path: str = "config.yaml") -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def load_state(cfg: Dict[str, Any]) -> Dict[str, Any]:
    """Load posting state from file"""
    state_file = cfg.get("state_file", "./posting_state.json")
    try:
        with open(state_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        # Initialize state for all models
        models = cfg.get("models", [])
        return {
            "current_cycle": 1,
            "model_posts": {model["name"]: 0 for model in models},
            "last_run": None
        }

def save_state(cfg: Dict[str, Any], state: Dict[str, Any]):
    """Save posting state to file"""
    state_file = cfg.get("state_file", "./posting_state.json")
    state["last_run"] = datetime.utcnow().isoformat() + "Z"
    with open(state_file, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)

def get_next_posts_for_cycle(cfg: Dict[str, Any], state: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Get the next posts for current cycle (one per model)"""
    models = cfg.get("models", [])
    cycle_posts = []
    
    for model in models:
        model_name = model["name"]
        current_post_num = state["model_posts"].get(model_name, 0)
        total_posts = model.get("total_posts", 8)
        
        # Check if this model has more posts to make
        if current_post_num >= total_posts:
            log(f"# [SKIP] {model_name} completed all {total_posts} posts")
            continue
            
        # Get the next post for this model
        next_post_num = current_post_num + 1
        pattern = model.get("pattern", ["free"])
        post_type = pattern[current_post_num] if current_post_num < len(pattern) else "free"
        
        # Build post configuration
        folder_name = model.get("folders", {}).get(post_type, "lingerie foto")
        media_config = model.get("media", {}).get(post_type, {"count": 1, "type": "photo"})
        
        post_config = {
            "name": f"post_{next_post_num}_{post_type}",
            "folder_filter": {"name_contains": folder_name},
            "media": media_config,
            "post_type": "free" if post_type == "free" else "paid"
        }
        
        # Add paid configuration if needed
        if post_type == "fans":
            post_config["paid"] = {"variant": "fans_only"}
        elif post_type == "paid":
            paid_settings = model.get("paid_settings", {}).get("paid", {})
            post_config["paid"] = paid_settings
            
        cycle_posts.append({
            "model": model,
            "post": post_config,
            "post_number": next_post_num,
            "post_type": post_type
        })
    
    return cycle_posts


# ---------------------------- Timezone & Pacing ----------------------------

def _tz(cfg: Dict[str, Any]) -> ZoneInfo:
    tzname = (cfg.get("timezone") or "UTC").strip()
    try:
        return ZoneInfo(tzname)
    except Exception:
        return ZoneInfo("UTC")

def _pick_delay_seconds(block: Dict[str, Any]) -> int:
    mode = (block.get("mode") or "none").lower()
    if mode == "fixed":
        return int(block.get("fixed_seconds", 0))
    if mode == "random":
        mn = int(block.get("random_min_seconds", 0))
        mx = int(block.get("random_max_seconds", mn))
        if mx < mn:
            mx = mn
        return random.randint(mn, mx)
    return 0

async def sleep_seconds(sec: int, label: str):
    if sec <= 0:
        return
    log(f"⏳ Waiting {sec} seconds ({label})...")
    end = time.time() + sec
    while True:
        rem = end - time.time()
        if rem <= 0:
            break
        await asyncio.sleep(min(5, rem))

def _parse_time_hhmm(hhmm: str, tz: ZoneInfo) -> Optional[datetime]:
    try:
        now = datetime.now(tz)
        H, M = [int(x) for x in hhmm.split(":")]
        target = now.replace(hour=H, minute=M, second=0, microsecond=0)
        if target <= now:
            return None
        return target
    except Exception:
        return None

async def maybe_wait_for_post_time(post: Dict[str, Any], cfg: Dict[str, Any]):
    when = (post.get("schedule") or {}).get("at")
    if not when:
        return
    tz = _tz(cfg)
    target = _parse_time_hhmm(str(when), tz)
    if not target:
        log(f'🕒 schedule.at="{when}" already passed or invalid; running now')
        return
    delta = int((target - datetime.now(tz)).total_seconds())
    await sleep_seconds(max(0, delta), f'wait until {when}')


# ---------------------------- Browser lifecycle (per post) ----------------------------

async def load_cookies(context: BrowserContext, cookies_path: str):
    p = Path(cookies_path)
    if not p.exists():
        log(f"No cookies at {cookies_path}; proceeding unauthenticated")
        return
    try:
        cookies = json.loads(p.read_text(encoding="utf-8"))
        await context.add_cookies(cookies)
        log(f"Loaded cookies from {cookies_path}")
    except Exception as e:
        log(f"Failed to load cookies: {e}")

async def open_creators_page(cfg: Dict[str, Any]):
    pw = await async_playwright().start()
    browser: Browser = await pw.chromium.launch(headless=cfg.get("headless", True))
    context: BrowserContext = await browser.new_context()
    await load_cookies(context, cfg["cookies_path"])
    page: Page = await context.new_page()
    log("# [NAV] Go to creators page")
    await page.goto("https://f2f.com/agency/creators/", wait_until="domcontentloaded", timeout=cfg["timeouts"]["long_ms"])
    
    # Handle cookie popup if it appears
    try:
        log("# [BTN] Check for cookie popup")
        cookie_accept = page.locator('button:has-text("Accept"), button:has-text("Accepteren"), button:has-text("OK"), [data-testid="cookie-accept"], .cookie-accept')
        await cookie_accept.wait_for(state="visible", timeout=3000)
        await cookie_accept.click()
        log("# [BTN] Clicked accept cookies")
        await asyncio.sleep(5)  # Wait 5 seconds after accepting cookies
    except Exception:
        log("# [INFO] No cookie popup found or already accepted")
    
    return pw, browser, context, page

async def close_browser(pw, browser: Browser, context: BrowserContext):
    try:
        await context.storage_state(path="storage_state.json")
    except Exception:
        pass
    await browser.close()
    await pw.stop()


# ---------------------------- Generic helpers ----------------------------

def _clamp_price(v: Any) -> float:
    try:
        price = float(v)
        return max(5.0, min(999.0, price))  # Min 5.00, Max 999.00
    except Exception:
        return 5.0

async def safe_click(locator, retries: int = 2, delay: float = 0.6, button_name: str = "Unknown Button"):
    """Enhanced safe_click with detailed error reporting"""
    last_error = None
    selector_info = "Unknown selector"
    
    try:
        # Try to get selector information for better error reporting
        selector_info = str(locator)
    except:
        pass
    
    log(f"# [BTN] Attempting to click: {button_name} (selector: {selector_info})")
    
    for attempt in range(retries + 1):
        try:
            log(f"# [BTN] Click attempt {attempt + 1}/{retries + 1} for {button_name}")
            await locator.wait_for(state="visible")
            await locator.click()
            await asyncio.sleep(max(3.0, 3.0))  # Minimum 3 seconds after each click
            log(f"# [SUCCESS] Successfully clicked: {button_name}")
            return
        except Exception as e:
            last_error = e
            log(f"# [ERROR] Click attempt {attempt + 1} failed for {button_name}: {str(e)}")
            if attempt < retries:
                log(f"# [RETRY] Waiting {max(delay, 3.0)} seconds before retry...")
                await asyncio.sleep(max(delay, 3.0))  # Minimum 3 seconds between retries too
    
    # If we get here, all attempts failed
    log(f"# [FATAL] All {retries + 1} click attempts failed for {button_name}")
    log(f"# [FATAL] Final error: {str(last_error)}")
    log(f"# [FATAL] Selector: {selector_info}")
    raise last_error

async def wheel_scroll(page: Page, dy: int = 1200, sleep: float = 0.5):
    await page.mouse.wheel(0, dy)
    await asyncio.sleep(sleep)

async def screenshot_on_error(page: Page, prefix: str = "error") -> str:
    Path("logs/screens").mkdir(parents=True, exist_ok=True)
    fp = f"logs/screens/{prefix}_{int(time.time())}.png"
    try:
        await page.screenshot(path=fp, full_page=True)
    except Exception:
        pass
    return fp

async def click_next_in_popup(page: Page, cfg: Dict[str, Any], label: str = "Next"):
    log(f'# [BTN] {label} (popup)')
    btn = page.locator(f'div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("{label}")')
    await safe_click(btn, retries=cfg["retries"]["clicks"], button_name=f"{label} (popup)")

async def click_publish_now(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] Publish now')
    btn = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Publish now")')
    await safe_click(btn, retries=cfg["retries"]["clicks"], button_name="Publish now")

async def clear_and_type_price(input_locator, price: float):
    await input_locator.click()
    try:
        await input_locator.press("Control+a")
    except Exception:
        pass
    await input_locator.press("Backspace")
    await input_locator.type(f"{price:.2f}")


# ---------------------------- Excel helpers (Caption MOVE) ----------------------------

def _ensure_wb_ws(path: str, sheet: str):
    p = Path(path)
    if p.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
    return wb, ws

def _read_column_values(ws, col_letter: str) -> List[str]:
    col_idx = column_index_from_string(col_letter)
    vals = []
    # Skip row 1 (header) - start from row 2
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v is not None and str(v).strip() != "":
            vals.append(str(v))
    return vals

def _write_column_values(ws, col_letter: str, values: List[str]):
    # clear entire column from row 2 onwards (preserve header in row 1)
    col_idx = column_index_from_string(col_letter)
    for row in range(2, max(ws.max_row, len(values)) + 50):  # generous clear
        ws.cell(row=row, column=col_idx).value = None
    # Write values starting from row 2 (preserve header)
    for i, text in enumerate(values, start=2):
        ws.cell(row=i, column=col_idx).value = text

def pop_caption_from_excel(excel_cfg: Dict[str, Any], strategy: str, model_name: str = None) -> str:
    wb, ws = _ensure_wb_ws(excel_cfg["path"], excel_cfg.get("sheet", "Sheet1"))
    col = excel_cfg.get("column", "A")
    values = _read_column_values(ws, col)
    if not values:
        raise RuntimeError("Source caption Excel is empty.")
    if strategy.lower() == "random":
        idx = random.randrange(0, len(values))
    else:
        idx = 0
    chosen = values[idx]
    remaining = [v for i, v in enumerate(values) if i != idx]
    _write_column_values(ws, col, remaining)
    wb.save(excel_cfg["path"])
    return chosen

def archive_used_caption(caption: str, archive_cfg: Dict[str, Any], model: str, post: str):
    typ = (archive_cfg.get("type") or "json").lower()
    if typ == "excel":
        ex = archive_cfg.get("excel", {})
        wb, ws = _ensure_wb_ws(ex["path"], ex.get("sheet", "Used"))
        col = ex.get("column", "A")
        existing = _read_column_values(ws, col)
        existing.append(caption)
        _write_column_values(ws, col, existing)
        wb.save(ex["path"])
        return
    # json default
    path = archive_cfg.get("json_path", "./captions/used_global.json")
    p = Path(path)
    data = {"captions": []}
    if p.exists():
        try:
            data = json.loads(p.read_text(encoding="utf-8")) or {"captions": []}
        except Exception:
            data = {"captions": []}
    data["captions"].append({
        "caption": caption,
        "model": model,
        "post": post,
        "ts": datetime.utcnow().isoformat() + "Z"
    })
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def replenish_excel_if_empty(excel_cfg: Dict[str, Any], archive_cfg: Dict[str, Any]) -> bool:
    """Return True if replenished."""
    wb, ws = _ensure_wb_ws(excel_cfg["path"], excel_cfg.get("sheet", "Sheet1"))
    col = excel_cfg.get("column", "A")
    current = _read_column_values(ws, col)
    if current:
        return False
    if not archive_cfg.get("replenish_when_empty", True):
        return False

    typ = (archive_cfg.get("type") or "json").lower()
    new_vals: List[str] = []

    if typ == "excel":
        ex = archive_cfg.get("excel", {})
        awb, aws = _ensure_wb_ws(ex["path"], ex.get("sheet", "Used"))
        arch_col = ex.get("column", "A")
        new_vals = _read_column_values(aws, arch_col)
    else:
        path = archive_cfg.get("json_path", "./captions/used_global.json")
        p = Path(path)
        if p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8")) or {"captions": []}
                new_vals = [c.get("caption", "") for c in data.get("captions", []) if c.get("caption")]
                if archive_cfg.get("after_replenish_clear_archive", False):
                    p.write_text(json.dumps({"captions": []}, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                new_vals = []

    if not new_vals:
        return False

    _write_column_values(ws, col, new_vals)
    wb.save(excel_cfg["path"])
    log(f"♻️ Replenished source captions from archive ({len(new_vals)} items).")
    return True


# ---------------------------- Media helpers (user-provided) ----------------------------

async def scroll_media_grid(page: Page) -> bool:
    # [STEP] Scroll media grid to load more items
    print("⏳ Scrolling inside media grid to load more items...")
    grid_selector = 'div.ZjJRiW_gridContainer.abvcKa_grid > div'
    last_count = 0
    attempts_without_new = 0
    max_attempts_without_new = 5
    while attempts_without_new < max_attempts_without_new:
        try:
            await page.wait_for_selector(grid_selector, timeout=30000)
            media_items = await page.query_selector_all(grid_selector)
            current_count = len(media_items)
            if current_count > last_count:
                print(f"📸 Media count increased: {last_count} → {current_count}")
                last_count = current_count
                attempts_without_new = 0
                if media_items:
                    await media_items[-1].scroll_into_view_if_needed()
                    await asyncio.sleep(3)
            else:
                attempts_without_new += 1
                print(f"⚠️ No new media (Try {attempts_without_new}/{max_attempts_without_new})")
                if media_items:
                    await media_items[-1].scroll_into_view_if_needed()
                    await asyncio.sleep(2)
        except Exception as e:
            print(f"⚠️ Error scrolling media grid: {e}")
            break
    print(f"✅ Finished scrolling, found {last_count} media items")
    return last_count > 0

async def select_random_media(page: Page, media_type: str = "photo", count: int = 1) -> bool:
    print(f"🎲 Selecting {count} random {media_type} item(s)...")
    grid_selector = 'div.ZjJRiW_gridContainer.abvcKa_grid > div'
    try:
        await page.wait_for_selector(grid_selector, timeout=45000)
        media_items = await page.query_selector_all(grid_selector)
        if len(media_items) < count:
            print(f"❌ Not enough media items found. Need {count}, found {len(media_items)}")
            return False
        selected_items = random.sample(media_items, count)
        for i, item in enumerate(selected_items):
            await item.click()
            print(f"✅ Selected {media_type} item {i + 1}/{count}")
            await asyncio.sleep(1)
        print(f"✅ Selected {count} random {media_type} item(s)")
        return True
    except Exception as e:
        print(f"❌ Error selecting random media: {e}")
        return False

async def select_random_video(page: Page, media_type: str = "video", count: int = 1) -> bool:
    print(f"🎬 Selecting {count} random {media_type}(s)...")
    grid_selector = 'div.ZjJRiW_gridContainer.abvcKa_grid > div'
    try:
        await page.wait_for_selector(grid_selector, timeout=45000)
        media_items = await page.query_selector_all(grid_selector)
        if len(media_items) < count:
            print(f"❌ Not enough video items found. Need {count}, found {len(media_items)}")
            return False
        selected_items = random.sample(media_items, count)
        for i, item in enumerate(selected_items):
            await item.click()
            print(f"✅ Selected {media_type} item {i + 1}/{count}")
            await asyncio.sleep(1)
        print(f"✅ Selected {count} random {media_type}(s)")
        return True
    except Exception as e:
        print(f"❌ Error selecting random video: {e}")
        return False


# ---------------------------- Core steps ----------------------------

async def search_and_open_creator(page: Page, cfg: Dict[str, Any], model: Dict[str, Any]):
    log(f'# [STEP] Search creator: {model["name"]}')
    # [FIELD] Search creator
    search = page.locator('input.iLRJUq_input[name="q"][placeholder="Search creator"]')
    await search.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await search.click()
    await asyncio.sleep(1.0)  # Wait after clicking search field
    await search.fill(model["name"])
    await asyncio.sleep(2.0)  # Wait after typing
    # Click the only result
    result = page.locator(f'a:has-text("{model["name"]}")').first
    if not await result.count():
        result = page.locator(f'button:has-text("{model["name"]}"), div:has-text("{model["name"]}")').first
    await result.wait_for(state="visible", timeout=cfg["timeouts"]["long_ms"])
    await safe_click(result, retries=cfg["retries"]["clicks"], button_name=f"Creator search result: {model['name']}")

async def click_create_then_post(page: Page, cfg: Dict[str, Any]):
    log("# [BTN] Click Create (not Messages)")
    create_btn = page.locator('div.ywtmnG_buttonsContainer >> button._7lrYkG_button._7lrYkG_tertiary:has(span._7lrYkG_text:has-text("Create"))')
    await safe_click(create_btn, retries=cfg["retries"]["clicks"], button_name="Create button")
    log('# [BTN] Click "Post"')
    post_item = page.locator('a.x9qUeG_linkItem:has(.x9qUeG_name:has-text("Post"))')
    await safe_click(post_item, retries=cfg["retries"]["clicks"], button_name="Post menu item")

async def click_parent_folder(page: Page, parent_folder_name: str) -> bool:
    """Click the parent folder first (e.g., 'software') before accessing nested folders"""
    try:
        log(f"# [BTN] Clicking parent folder: {parent_folder_name}")
        
        # Look for parent folder in the folder structure
        folder_containers = await page.query_selector_all("div.OsMxUq_folder")
        
        for container in folder_containers:
            try:
                # Get folder name from p.OsMxUq_name
                name_elem = await container.query_selector("p.OsMxUq_name")
                if not name_elem:
                    continue
                
                folder_name = (await name_elem.inner_text()).lower().strip()
                
                if parent_folder_name.lower() in folder_name:
                    # Get the clickable link
                    link_elem = await container.query_selector("a.OsMxUq_folderImage")
                    if link_elem:
                        await link_elem.click()
                        log(f"# [SUCCESS] Clicked parent folder: {folder_name}")
                        return True
                    else:
                        await container.click()
                        log(f"# [SUCCESS] Clicked parent folder container: {folder_name}")
                        return True
                        
            except Exception as e:
                log(f"# [DEBUG] Error checking folder container: {e}")
                continue
                
        log(f"# [ERROR] Parent folder '{parent_folder_name}' not found")
        return False
        
    except Exception as e:
        log(f"# [ERROR] Error clicking parent folder: {e}")
        return False

async def select_folder(page: Page, cfg: Dict[str, Any], post: Dict[str, Any], model: Dict[str, Any]):
    filt = post.get("folder_filter", {}) or {}
    href_need = (filt.get("href_contains") or "").lower()
    name_need = (filt.get("name_contains") or "").lower()
    
    # Check if this model has nested folders (optional - only for specific clients like Ryan)
    nested_config = model.get("nested_folders", {})
    parent_folder = nested_config.get("parent") if nested_config else None
    
    if parent_folder:
        log(f"# [STEP] Nested folder structure detected - clicking parent folder '{parent_folder}' first")
        
        # First click the parent folder
        await asyncio.sleep(3.0)
        parent_found = await click_parent_folder(page, parent_folder)
        if not parent_found:
            log(f"# [WARNING] Parent folder '{parent_folder}' not found, trying direct folder access")
        else:
            # Wait for nested folders to load
            await asyncio.sleep(10.0)  # Longer wait for nested structure
            log(f"# [INFO] Parent folder clicked, waiting for nested folders to load...")
    
    log(f"# [STEP] Select folder by name: looking for '{name_need}'")
    
    # Wait for media library to load
    await asyncio.sleep(5.0)

    async def try_find():
        # Look for F2F folder structure: div.OsMxUq_folder containing p.OsMxUq_name
        try:
            folder_containers = await page.query_selector_all("div.OsMxUq_folder")
            available_folders = []
            
            for container in folder_containers:
                try:
                    # Get folder name from p.OsMxUq_name
                    name_elem = await container.query_selector("p.OsMxUq_name")
                    if not name_elem:
                        continue
                    
                    folder_name = (await name_elem.inner_text()).lower().strip()
                    if folder_name:
                        available_folders.append(folder_name)
                    
                    # Check if this matches our filter (no need for model matching since we're already on the model's page)
                    if name_need in folder_name:
                        # Get the clickable link (a.OsMxUq_folderImage)
                        link_elem = await container.query_selector("a.OsMxUq_folderImage")
                        if link_elem:
                            href = await link_elem.get_attribute("href") or ""
                            log(f"# [FOUND] Matched folder: '{folder_name}' (href: {href})")
                            return link_elem
                        else:
                            # Fallback: try to click the container itself
                            log(f"# [FOUND] Matched folder: '{folder_name}' (clicking container)")
                            return container
                            
                except Exception as e:
                    log(f"# [DEBUG] Error processing folder container: {e}")
                    continue
                    
        except Exception as e:
            log(f"# [DEBUG] Error finding folder containers: {e}")
        
        # Log available folders for debugging (remove duplicates)
        unique_folders = list(set(available_folders))
        if unique_folders:
            log(f"# [DEBUG] Available media folders found: {unique_folders[:15]}")
        else:
            log("# [DEBUG] No folders found with current selectors, checking page content...")
            # Try to get all text content for debugging
            try:
                page_text = await page.inner_text("body")
                if "lingerie" in page_text.lower() or "foto" in page_text.lower():
                    log("# [DEBUG] Found 'lingerie' or 'foto' in page text - folders might be loading")
                else:
                    log("# [DEBUG] No folder-related text found on page")
            except Exception:
                log("# [DEBUG] Could not read page text")
        
        return None

    # Try to find the folder (no scrolling to avoid bot detection)
    found = await try_find()
    if found:
        await asyncio.sleep(1.0)
        await found.click()
        await asyncio.sleep(10.0)  # Wait 10 seconds for folder to load properly
        log(f"# [SUCCESS] Selected folder containing '{name_need}'")
        return
    
    raise RuntimeError(f"Folder not found with provided filters. Looking for: '{name_need}'. Check available folders in debug output above.")

async def select_media_step(page: Page, cfg: Dict[str, Any], post: Dict[str, Any]):
    n = int(post["media"]["count"])
    media_type = (post["media"].get("type") or "photo").lower()
    log(f"# [STEP] Load & select {n} media items (type={media_type})")
    await scroll_media_grid(page)
    ok = await (select_random_video if media_type == "video" else select_random_media)(page, media_type, n)
    if not ok:
        raise RuntimeError("Failed to select the requested number of media items")

async def click_next_library(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] Next (library)')
    next_btn = page.locator('div.abvcKa_desktopButtons >> div._8HzuwW_button._8HzuwW_primary:has-text("Next")')
    await safe_click(next_btn, retries=cfg["retries"]["clicks"], button_name="Next (library)")
    await asyncio.sleep(10.0)  # Wait 10 seconds after clicking Next (library)

async def maybe_add_caption(page: Page, cfg: Dict[str, Any], model: Dict[str, Any], post: Dict[str, Any]):
    cap_cfg = (post.get("caption") if "caption" in post else model.get("caption")) or {}
    if not cap_cfg.get("enabled", False):
        log("# [STEP] Caption disabled; skipping")
        return

    # 1) Replenish if empty
    archive_cfg = (post.get("caption_archive") or model.get("caption_archive") or cfg.get("caption_archive") or {})
    if replenish_excel_if_empty(cap_cfg["excel"], archive_cfg):
        log("# [STEP] Replenished captions into source Excel")

    # 2) Pop (remove from source)
    pick_strategy = cap_cfg["excel"].get("pick_strategy", "sequential")
    caption = pop_caption_from_excel(cap_cfg["excel"], pick_strategy)

    log('# [BTN] Click "Edit"')
    edit_btn = page.locator('button._63d-OW_actionButton:has(span._63d-OW_text:text-is("Edit"))')
    await safe_click(edit_btn, retries=cfg["retries"]["clicks"], button_name="Edit (caption)")

    log("# [FIELD] Caption textarea")
    ta = page.locator("textarea.n3Mc_G_mentions__input")
    await ta.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await ta.fill(caption)
    await asyncio.sleep(1.0)  # Wait after filling caption

    hashtags_cfg = cap_cfg.get("hashtags", {})
    if hashtags_cfg.get("enabled", False):
        await page.keyboard.press("Enter")
        await page.keyboard.press("Enter")
        tags = hashtags_cfg.get("list", [])
        if tags:
            await ta.type(" ".join(tags))
            await ta.type(" ")  # ensure last hashtag registers
            await asyncio.sleep(1.0)  # Wait after adding hashtags

    log('# [BTN] Click "Save"')
    save_btn = page.locator('button._63d-OW_actionButton:has(span._63d-OW_text:text-is("Save"))')
    await safe_click(save_btn, retries=cfg["retries"]["clicks"], button_name="Save (caption)")

    # 3) Archive used caption
    archive_used_caption(caption, archive_cfg, model["name"], post.get("name", "post"))

async def click_next_composer(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] Next (composer)')
    btn = page.locator('button.L6JeHW_nextButton._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn, retries=cfg["retries"]["clicks"], button_name="Next (composer)")

async def free_flow_publish_now(page: Page, cfg: Dict[str, Any]):
    await click_next_in_popup(page, cfg, "Next")
    await click_publish_now(page, cfg)


# ---------------------------- Paid flow ----------------------------

async def switch_all_free_to_paid(page: Page, cfg: Dict[str, Any], media_count: int = None):
    log('# [STEP] Toggle all "Free" chips to paid')
    
    # If we don't know media count, try to detect it
    if media_count is None:
        # Try to count selected media items or use a reasonable default
        media_count = 5  # Default fallback
    
    log(f'# [INFO] Clicking Free for {media_count} media items individually')
    
    # Multiple selectors to try for free buttons
    free_selectors = [
        'div.access-status span.free',
        'span.free',
        'div.access-status',
        'div[style*="background-color: rgba(0, 0, 0, 0.15)"]',
        'button:has-text("Free")',
        'div[class*="free"]',
        'span[class*="free"]'
    ]
    
    # Click "Free" for each media item
    for photo_index in range(media_count):
        log(f'# [BTN] Clicking Free for media item {photo_index + 1}/{media_count}')
        
        free_clicked = False
        for selector in free_selectors:
            try:
                # Wait for page to stabilize
                await asyncio.sleep(1.0)
                
                # Find all free elements and click the first available one
                free_elements = await page.query_selector_all(selector)
                if len(free_elements) > 0:
                    # Click the first available free element
                    await free_elements[0].click()
                    log(f'# [SUCCESS] Clicked Free for media {photo_index + 1} using selector: {selector}')
                    free_clicked = True
                    await asyncio.sleep(3.0)  # Wait for page to update
                    break
            except Exception as e:
                log(f'# [DEBUG] Selector {selector} failed: {e}')
                continue
        
        if not free_clicked:
            log(f'# [WARNING] Could not click Free for media item {photo_index + 1}')
            # Continue anyway, might still work
    
    log(f'# [STEP] Completed Free button clicking for {media_count} media items')

async def paid_variant_fans_only(page: Page, cfg: Dict[str, Any]):
    log('# [OPT] Fans only -> Next')
    await click_next_in_popup(page, cfg, "Next")

async def paid_variant_non_fans_pay(page: Page, cfg: Dict[str, Any], price_followers: float):
    log('# [OPT] Only non-fans must pay')
    card = page.locator('div[data-testid="expanding-button-card"]:has(.ztm8ka_header:has-text("Only non-fans must pay"))')
    await safe_click(card, retries=cfg["retries"]["clicks"], button_name="Only non-fans must pay card")
    log('# [FIELD] Followers price')
    field = page.locator('div.KKpREq_inputWrapper input[data-testid="decimal-visual-input"]').first
    await field.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await clear_and_type_price(field, _clamp_price(price_followers))
    await click_next_in_popup(page, cfg, "Next")

async def paid_variant_paid_for_everyone(page: Page, cfg: Dict[str, Any], price_followers: float, price_fans: float):
    log('# [OPT] Paid for everyone')
    card = page.locator('div[data-testid="expanding-button-card"]:has(.ztm8ka_header:has-text("Paid for everyone"))')
    await safe_click(card, retries=cfg["retries"]["clicks"], button_name="Paid for everyone card")
    inputs = page.locator('input[data-testid="decimal-visual-input"]')
    await inputs.first.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    log('# [FIELD] Followers price')
    await clear_and_type_price(inputs.nth(0), _clamp_price(price_followers))
    log('# [FIELD] Fans price')
    await clear_and_type_price(inputs.nth(1), _clamp_price(price_fans))
    await click_next_in_popup(page, cfg, "Next")

async def paid_variant_vip_post(page: Page, cfg: Dict[str, Any], price_fans: float):
    log('# [OPT] VIP post (Only paid available for fans)')
    card = page.locator('div[data-testid="expanding-button-card"]:has(.ztm8ka_header:has-text("VIP post"))')
    await safe_click(card, retries=cfg["retries"]["clicks"], button_name="VIP post card")
    log('# [FIELD] Fans price')
    vip_field = card.locator('input[data-testid="decimal-visual-input"]').first
    await vip_field.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await clear_and_type_price(vip_field, _clamp_price(price_fans))
    await click_next_in_popup(page, cfg, "Next")

async def paid_flow(page: Page, cfg: Dict[str, Any], post: Dict[str, Any]):
    media_count = int(post["media"]["count"])
    await switch_all_free_to_paid(page, cfg, media_count)
    await click_next_in_popup(page, cfg, "Next")
    paid_cfg = post.get("paid", {}) or {}
    variant = (paid_cfg.get("variant") or "").strip().lower()
    prices = paid_cfg.get("prices", {}) or {}
    followers_p = prices.get("followers")
    fans_p = prices.get("fans")
    if variant == "fans_only":
        await paid_variant_fans_only(page, cfg)
    elif variant == "non_fans_pay":
        if followers_p is None:
            raise RuntimeError("Missing prices.followers for 'non_fans_pay'")
        await paid_variant_non_fans_pay(page, cfg, followers_p)
    elif variant == "paid_for_everyone":
        if followers_p is None or fans_p is None:
            raise RuntimeError("Missing prices.followers or prices.fans for 'paid_for_everyone'")
        await paid_variant_paid_for_everyone(page, cfg, followers_p, fans_p)
    elif variant == "vip_post":
        if fans_p is None:
            raise RuntimeError("Missing prices.fans for 'vip_post'")
        await paid_variant_vip_post(page, cfg, fans_p)
    else:
        raise RuntimeError(f"Unknown paid.variant: {variant}")
    await click_publish_now(page, cfg)


# ---------------------------- Orchestration ----------------------------

# Removed _normalize_posts_for_model - using rotation logic instead

async def run_single_post(model: Dict[str, Any], post: Dict[str, Any], cfg: Dict[str, Any]):
    await maybe_wait_for_post_time(post, cfg)

    pw, browser, context, page = await open_creators_page(cfg)
    try:
        await search_and_open_creator(page, cfg, model)
        await click_create_then_post(page, cfg)
        await select_folder(page, cfg, post, model)
        await select_media_step(page, cfg, post)
        await click_next_library(page, cfg)
        await maybe_add_caption(page, cfg, model, post)
        await click_next_composer(page, cfg)

        post_type = (post.get("post_type") or model.get("post_type") or "free").lower()
        if post_type == "paid":
            log("# [FLOW] Paid path selected")
            await paid_flow(page, cfg, post)
        else:
            log("# [FLOW] Free path selected")
            await free_flow_publish_now(page, cfg)

        # Return to creators page before closing (safer)
        await page.goto("https://f2f.com/agency/creators/", wait_until="domcontentloaded")

    except Exception as e:
        fp = await screenshot_on_error(page, prefix=f'{model["name"]}_{post.get("name","post")}')
        log(f"❌ Error: {e}\n{traceback.format_exc()}\nScreenshot: {fp}")
        raise
    finally:
        await close_browser(pw, browser, context)

def _normalize_posts_for_model(model: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Convert model pattern to individual posts (for sequential mode)"""
    pattern = model.get("pattern", ["free"])
    total_posts = model.get("total_posts", len(pattern))
    posts = []
    
    for i in range(total_posts):
        post_type = pattern[i] if i < len(pattern) else "free"
        folder_name = model.get("folders", {}).get(post_type, "lingerie foto")
        media_config = model.get("media", {}).get(post_type, {"count": 1, "type": "photo"})
        
        post = {
            "name": f"post_{i+1}_{post_type}",
            "folder_filter": {"name_contains": folder_name},
            "media": media_config,
            "post_type": "free" if post_type == "free" else "paid"
        }
        
        # Add paid configuration if needed
        if post_type == "fans":
            post["paid"] = {"variant": "fans_only"}
        elif post_type == "paid":
            paid_settings = model.get("paid_settings", {}).get("paid", {})
            post["paid"] = paid_settings
            
        posts.append(post)
    
    return posts

async def main():
    cfg = load_cfg()
    posting_mode = cfg.get("posting_mode", "sequential").lower()
    
    log("=" * 72)
    log(f"F2F POSTING BOT - {posting_mode.upper()} MODE")
    log("=" * 72)
    
    if posting_mode == "rotation":
        await run_rotation_mode(cfg)
    else:
        await run_sequential_mode(cfg)

async def run_rotation_mode(cfg: Dict[str, Any]):
    """Run in rotation mode: auto-complete all cycles until done"""
    state = load_state(cfg)
    max_cycles = 50  # Safety limit
    cycle_delay = cfg.get("pace", {}).get("cycle_delay", 1800)  # 30 min between cycles
    
    log(f"🔄 ROTATION MODE - Starting from Cycle: {state['current_cycle']}")
    log("🤖 Will auto-complete ALL cycles until finished!")
    
    while state['current_cycle'] <= max_cycles:
        log("=" * 72)
        log(f"🔄 CYCLE {state['current_cycle']}")
        log("=" * 72)
        
        # Get next posts for this cycle (one per model)
        cycle_posts = get_next_posts_for_cycle(cfg, state)
        
        if not cycle_posts:
            log("🎉 ALL MODELS COMPLETED! No more posts to make.")
            break
        
        log(f"📋 Cycle {state['current_cycle']}: {len(cycle_posts)} models to post")
        
        between_models_delay = cfg.get("pace", {}).get("between_models", 1200)
        
        for i, cycle_post in enumerate(cycle_posts, start=1):
            model = cycle_post["model"]
            post = cycle_post["post"]
            post_number = cycle_post["post_number"]
            post_type = cycle_post["post_type"]
            
            log("=" * 50)
            log(f"MODEL {i}/{len(cycle_posts)}: {model['name']} - Post {post_number} ({post_type.upper()})")
            log("=" * 50)
            
            try:
                await run_single_post(model, post, cfg)
                
                # Update state - increment this model's post count
                state["model_posts"][model["name"]] = post_number
                save_state(cfg, state)
                
                log(f"✅ Completed {model['name']} post {post_number}")
                
            except Exception as e:
                log(f"❌ Failed {model['name']} post {post_number}: {e}")
                # Continue with other models even if one fails
                
            # Wait between models (except last one)
            if i < len(cycle_posts):
                await sleep_seconds(between_models_delay, f"before next model ({between_models_delay//60} min)")
        
        # Update cycle number
        state["current_cycle"] += 1
        save_state(cfg, state)
        
        log("=" * 72)
        log(f"🎉 CYCLE {state['current_cycle']-1} COMPLETED!")
        log(f"📊 Current progress: {state['model_posts']}")
        
        # Check if all models are done
        models = cfg.get("models", [])
        all_done = True
        for model in models:
            model_name = model["name"]
            completed = state["model_posts"].get(model_name, 0)
            total = model.get("total_posts", 8)
            if completed < total:
                all_done = False
                log(f"📋 {model_name}: {completed}/{total} posts completed")
            else:
                log(f"✅ {model_name}: ALL {total} posts completed")
        
        if all_done:
            log("🎊 ALL MODELS FULLY COMPLETED! Bot finished successfully.")
            break
        else:
            log(f"🔄 Starting next cycle in {cycle_delay//60} minutes...")
            await sleep_seconds(cycle_delay, f"between cycles")
    
    if state['current_cycle'] > max_cycles:
        log(f"⚠️ Reached maximum cycles ({max_cycles}). Stopping for safety.")
    
    log("=" * 72)

async def run_sequential_mode(cfg: Dict[str, Any]):
    """Run in sequential mode: complete all posts for each model before next"""
    models: List[Dict[str, Any]] = cfg.get("models", [])
    between_models_delay = cfg.get("pace", {}).get("between_models", 1200)
    between_posts = cfg.get("pace", {}).get("between_posts", {}) or {"mode": "none"}

    for m_i, model in enumerate(models, start=1):
        log("=" * 72)
        log(f"MODEL {m_i}/{len(models)}: {model['name']}")
        log("=" * 72)

        posts = _normalize_posts_for_model(model)
        for p_i, post in enumerate(posts, start=1):
            log(f"--- Post {p_i}/{len(posts)} :: {post.get('name')} ---")
            await run_single_post(model, post, cfg)

            # Delay BETWEEN POSTS (same model)
            if p_i < len(posts):
                sec = _pick_delay_seconds(between_posts)
                await sleep_seconds(sec, f"between posts for {model['name']}")

        # Delay BETWEEN MODELS
        if m_i < len(models):
            await sleep_seconds(between_models_delay, f"between models ({between_models_delay//60} min)")

if __name__ == "__main__":
    asyncio.run(main())
