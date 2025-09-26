# dm_main.py
# F2F Mass DM Bot — async Playwright
# Adds per-campaign scheduling (HH:MM local), MESSAGE MOVE workflow (Excel -> archive JSON/Excel -> replenish),
# and per-model/per-campaign overrides. Fresh browser per campaign. Heavy # [BTN]/[FIELD]/[STEP] comments.

import asyncio
import json
import random
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml
from zoneinfo import ZoneInfo
from playwright.async_api import async_playwright, Page, Browser, BrowserContext
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

# ========================= 0. LOG + CONFIG =========================

def log(msg: str) -> None:
    print(f"[F2F DM] {msg}", flush=True)

def load_cfg(path: str = "dm_config.yaml") -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

# ========================= 1. TIME / PACE =========================

def _tz(cfg: Dict[str, Any]) -> ZoneInfo:
    tzname = (cfg.get("timezone") or "UTC").strip()
    try: return ZoneInfo(tzname)
    except Exception: return ZoneInfo("UTC")

def _parse_hhmm_local(hhmm: str, tz: ZoneInfo) -> Optional[datetime]:
    try:
        now = datetime.now(tz)
        H, M = [int(x) for x in hhmm.split(":")]
        target = now.replace(hour=H, minute=M, second=0, microsecond=0)
        if target <= now:
            return None
        return target
    except Exception:
        return None

async def maybe_wait_for_campaign_time(campaign: Dict[str, Any], cfg: Dict[str, Any]):
    when = (campaign.get("schedule") or {}).get("at")
    if not when:
        return
    tz = _tz(cfg)
    target = _parse_hhmm_local(str(when), tz)
    if not target:
        log(f'🕒 schedule.at="{when}" already passed/invalid; running now')
        return
    secs = int((target - datetime.now(tz)).total_seconds())
    if secs > 0:
        log(f"⏳ Waiting {secs}s until {when} local ({tz.key})...")
        end = time.time() + secs
        while True:
            rem = end - time.time()
            if rem <= 0: break
            await asyncio.sleep(min(5, rem))

def _pick_delay_seconds(block: Dict[str, Any]) -> int:
    mode = (block.get("mode") or "none").lower()
    if mode == "fixed": return int(block.get("fixed_seconds", 0))
    if mode == "random":
        mn = int(block.get("random_min_seconds", 0))
        mx = int(block.get("random_max_seconds", mn))
        if mx < mn: mx = mn
        return random.randint(mn, mx)
    return 0

# ========================= 2. GENERIC HELPERS =========================

def _clamp_price(value: Any, mn: float, mx: float) -> float:
    try: v = float(value)
    except Exception: v = mn
    return min(max(v, mn), mx)

async def safe_click(locator, retries: int = 3, delay: float = 3.0):
    last = None
    for _ in range(retries + 1):
        try:
            await locator.wait_for(state="visible")
            await locator.click()
            await asyncio.sleep(3.0)  # Wait after clicking
            return
        except Exception as e:
            last = e
            await asyncio.sleep(delay)
    raise last

async def screenshot_on_error(page: Page, prefix: str = "error") -> str:
    Path("logs/screens").mkdir(parents=True, exist_ok=True)
    fp = f"logs/screens/{prefix}_{int(time.time())}.png"
    try:
        await page.screenshot(path=fp, full_page=True)
    except Exception:
        pass
    return fp

async def wheel_scroll(page: Page, dy: int = 1200, sleep: float = 0.5):
    await page.mouse.wheel(0, dy)
    await asyncio.sleep(sleep)

# ========================= 3. MESSAGE MOVE HELPERS (Excel <-> Archive) =========================

def _ensure_wb_ws(path: str, sheet: str):
    p = Path(path)
    if p.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # ensure first sheet has the right name (openpyxl default is "Sheet")
        if sheet not in wb.sheetnames:
            ws0 = wb.active
            ws0.title = sheet
    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)
    return wb, ws

def _read_column_values(path: str, sheet: str, col_letter: str) -> List[str]:
    if not Path(path).exists():
        return []
    wb = load_workbook(path)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    col_idx = column_index_from_string(col_letter)
    vals = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None and str(v).strip() != "":
            vals.append(str(v))
    return vals

def _write_column_values(ws, col_letter: str, values: List[str]):
    col_idx = column_index_from_string(col_letter)
    # clear generously
    for r in range(1, max(ws.max_row, len(values)) + 50):
        ws.cell(row=r, column=col_idx).value = None
    for i, text in enumerate(values, start=1):
        ws.cell(row=i, column=col_idx).value = text

def pop_message_from_excel(excel_cfg: Dict[str, Any]) -> str:
    """Removes a message from source Excel (first non-empty row) and returns it."""
    path = excel_cfg["path"]; sheet = excel_cfg.get("sheet", "Sheet1"); col = excel_cfg.get("column", "A")
    strategy = (excel_cfg.get("pick_strategy") or "sequential").lower()
    wb, ws = _ensure_wb_ws(path, sheet)
    values = _read_column_values(path, sheet, col)
    if not values:
        raise RuntimeError("Source DM Excel is empty.")
    if strategy == "random":
        idx = random.randrange(0, len(values))
    else:
        idx = 0
    chosen = values[idx]
    remaining = [v for i, v in enumerate(values) if i != idx]
    _write_column_values(ws, col, remaining)
    wb.save(path)
    return chosen

def archive_used_message(text: str, archive_cfg: Dict[str, Any], model: str, campaign: str, excel_path: str = None):
    typ = (archive_cfg.get("type") or "json").lower()
    
    # Determine archive path
    if typ == "excel":
        ex = archive_cfg.get("excel", {})
        apath = ex["path"]; asheet = ex.get("sheet", "Used"); acol = ex.get("column", "A")
        awb, aws = _ensure_wb_ws(apath, asheet)
        existing = _read_column_values(apath, asheet, acol)
        existing.append(text)
        _write_column_values(aws, acol, existing)
        awb.save(apath)
        return
    
    # JSON archive - simple logic
    if archive_cfg.get("per_file_archives", True) and excel_path:
        # One Excel file = One archive file (based on Excel filename)
        excel_name = Path(excel_path).stem  # Get filename without extension
        archive_dir = archive_cfg.get("archive_dir", "./dm_archive")
        jpath = f"{archive_dir}/{excel_name}_used.json"
    else:
        # Use single global archive for everything
        jpath = archive_cfg.get("json_path", "./dm_archive/used_global.json")
    
    p = Path(jpath); data = {"messages": []}
    if p.exists():
        try: data = json.loads(p.read_text(encoding="utf-8")) or {"messages": []}
        except Exception: data = {"messages": []}
    data["messages"].append({
        "text": text,
        "model": model,
        "campaign": campaign,
        "ts": datetime.utcnow().isoformat() + "Z"
    })
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def replenish_excel_if_empty(excel_cfg: Dict[str, Any], archive_cfg: Dict[str, Any]) -> bool:
    """If source Excel column empty and replenish enabled, refill from archive. Returns True if replenished."""
    path = excel_cfg["path"]; sheet = excel_cfg.get("sheet", "Sheet1"); col = excel_cfg.get("column", "A")
    current = _read_column_values(path, sheet, col)
    if current:
        return False
    if not (archive_cfg.get("replenish_when_empty", True)):
        return False

    typ = (archive_cfg.get("type") or "json").lower()
    new_vals: List[str] = []

    if typ == "excel":
        ex = archive_cfg.get("excel", {})
        apath = ex["path"]; asheet = ex.get("sheet", "Used"); acol = ex.get("column", "A")
        new_vals = _read_column_values(apath, asheet, acol)
    else:
        # Simple archive path logic
        if archive_cfg.get("per_file_archives", True):
            # One Excel file = One archive file (based on Excel filename)
            excel_name = Path(path).stem  # Get filename without extension
            archive_dir = archive_cfg.get("archive_dir", "./dm_archive")
            jpath = f"{archive_dir}/{excel_name}_used.json"
        else:
            # Use single global archive for everything
            jpath = archive_cfg.get("json_path", "./dm_archive/used_global.json")
        
        p = Path(jpath)
        if p.exists():
            try:
                data = json.loads(p.read_text(encoding="utf-8")) or {"messages": []}
                new_vals = [m.get("text", "") for m in data.get("messages", []) if m.get("text")]
                if archive_cfg.get("after_replenish_clear_archive", False):
                    p.write_text(json.dumps({"messages": []}, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                new_vals = []

    if not new_vals:
        return False

    wb, ws = _ensure_wb_ws(path, sheet)
    _write_column_values(ws, col, new_vals)
    wb.save(path)
    log(f"♻️ Replenished DM source Excel from archive ({len(new_vals)} items).")
    return True

def resolve_message_cfg(campaign: Dict[str, Any], model: Dict[str, Any], root_cfg: Dict[str, Any]) -> Dict[str, Any]:
    """Campaign overrides model overrides root."""
    return (campaign.get("message") or model.get("message") or root_cfg.get("message") or {})

def resolve_message_archive_cfg(campaign: Dict[str, Any], model: Dict[str, Any], root_cfg: Dict[str, Any]) -> Dict[str, Any]:
    return (campaign.get("message_archive") or model.get("message_archive") or root_cfg.get("message_archive") or {})

# ========================= 4. BROWSER + LOGIN =========================

async def load_cookies(context: BrowserContext, cookies_path: str):
    p = Path(cookies_path)
    if not p.exists():
        log(f"No cookies at {cookies_path}; proceeding unauthenticated")
        return
    try:
        cookies = json.loads(p.read_text(encoding="utf-8"))
        await context.add_cookies(cookies)
        log(f"Loaded cookies from {cookies_path}")
    except Exception as e:
        log(f"Failed to load cookies: {e}")

async def open_creators_page(cfg: Dict[str, Any]):
    pw = await async_playwright().start()
    browser: Browser = await pw.chromium.launch(headless=cfg.get("headless", True))
    context: BrowserContext = await browser.new_context()
    await load_cookies(context, cfg["cookies_path"])
    page: Page = await context.new_page()
    log("# [NAV] Go to creators page")
    await page.goto("https://f2f.com/agency/creators/", wait_until="domcontentloaded", timeout=cfg["timeouts"]["long_ms"])

    # Handle cookie popup if it appears
    try:
        log("# [BTN] Check for cookie popup")
        cookie_accept = page.locator('button:has-text("Accept"), button:has-text("Accepteren"), button:has-text("OK"), [data-testid="cookie-accept"], .cookie-accept')
        await cookie_accept.wait_for(state="visible", timeout=3000)
        await cookie_accept.click()
        log("# [BTN] Clicked accept cookies")
        await asyncio.sleep(5)  # Wait 5 seconds after accepting cookies
    except Exception:
        log("# [INFO] No cookie popup found or already accepted")

    return pw, browser, context, page

async def close_browser(pw, browser: Browser, context: BrowserContext):
    try:
        await context.storage_state(path="storage_state_dm.json")
    except Exception:
        pass
    await browser.close()
    await pw.stop()

# ========================= 5. SEARCH MODEL + OPEN MESSENGER =========================

async def search_and_open_creator(page: Page, cfg: Dict[str, Any], model: Dict[str, Any]):
    log(f'# [STEP] Search creator: {model["name"]}')
    # [FIELD] Search creator
    search = page.locator('input.iLRJUq_input[name="q"][placeholder="Search creator"]')
    await search.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await search.click()
    await asyncio.sleep(3.0)  # Wait after clicking search field
    await search.fill(model["name"])
    await asyncio.sleep(3.0)  # Wait after typing
    # Click the only result
    result = page.locator(f'a:has-text("{model["name"]}")').first
    if not await result.count():
        result = page.locator(f'button:has-text("{model["name"]}"), div:has-text("{model["name"]}")').first
    await result.wait_for(state="visible", timeout=cfg["timeouts"]["long_ms"])
    await safe_click(result, retries=cfg["retries"]["clicks"])

async def open_messages(page: Page, cfg: Dict[str, Any]):
    # [BTN] Messages button
    log('# [BTN] Open "Messages"')
    messages = page.locator('div.ywtmnG_buttonsContainer a:has(button:has-text("Messages"))')
    await safe_click(messages, retries=cfg["retries"]["clicks"])
    # [BTN] Mass message tab
    log('# [BTN] "Mass message" tab')
    mass_tab = page.locator('div.Rb86QW_headerContainer .Rb86QW_headerItems[data-testid="tab-header-1"]:has-text("Mass message")')
    await safe_click(mass_tab, retries=cfg["retries"]["clicks"])
    # [BTN] + create mass message
    log('# [BTN] Create mass message (+)')
    plus_btn = page.locator('button._6fMQbG_createMassButton[data-testid="create-mass-message-btn"]')
    await safe_click(plus_btn, retries=cfg["retries"]["clicks"])

# ========================= 6. AUDIENCE SELECTION =========================

async def audience_followers_only_next(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] Audience: Followers only -> Next')
    btn_next = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn_next, retries=cfg["retries"]["clicks"])

async def audience_fans_only_next(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] Audience: Fans only (toggle cards)')
    # Deselect Followers (click card)
    followers_card = page.locator('div.v-h2ZW_card[data-testid="checkbox-card"]:has-text("Followers")')
    await followers_card.click()
    # Select Fans
    fans_card = page.locator('div.v-h2ZW_card[data-testid="checkbox-card"]:has-text("Fans")')
    await safe_click(fans_card, retries=cfg["retries"]["clicks"])
    # Next
    btn_next = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn_next, retries=cfg["retries"]["clicks"])

async def audience_followers_and_fans_next(page: Page, cfg: Dict[str, Any]):
    log('# [BTN] Audience: Followers & Fans (add Fans)')
    fans_card = page.locator('div.v-h2ZW_card[data-testid="checkbox-card"]:has-text("Fans")')
    await safe_click(fans_card, retries=cfg["retries"]["clicks"])
    btn_next = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn_next, retries=cfg["retries"]["clicks"])

# ========================= 7. MESSAGE TYPE STEP =========================

async def type_textarea(page: Page, cfg: Dict[str, Any], text: str):
    log('# [FIELD] DM textarea')
    textarea = page.locator('div.KKpREq_inputWrapper textarea.KKpREq_textArea[name="message"]')
    await textarea.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await textarea.click()
    await asyncio.sleep(2.0)  # Wait after clicking textarea
    await textarea.fill(text or "")
    await asyncio.sleep(2.0)  # Wait after filling text

async def text_only_flow(page: Page, cfg: Dict[str, Any], text: str):
    log('# [BTN] Type: Text only -> Next')
    btn_next = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn_next, retries=cfg["retries"]["clicks"])
    await type_textarea(page, cfg, text)
    
    log('# [BTN] Create (text only)')
    btn_create = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Create")')
    await safe_click(btn_create, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(5.0)  # Wait 5 seconds after creating

async def request_tip_flow(page: Page, cfg: Dict[str, Any], text: str, amount: float):
    log('# [BTN] Type: Request tip')
    req_tip = page.locator('div.sqNJyW_card[data-testid="button-card"]:has(.sqNJyW_header:has-text("Request tip"))')
    await safe_click(req_tip, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(3.0)  # Wait for tip form to load
    
    log('# [BTN] Next (after selecting Request tip)')
    btn_next = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn_next, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(3.0)  # Wait for form to load
    
    await type_textarea(page, cfg, text)
    log('# [FIELD] Tip amount input')
    amount_input = page.locator('input.KKpREq_input[data-testid="decimal-visual-input"]')
    await amount_input.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
    await amount_input.click()
    await asyncio.sleep(1.0)  # Wait after clicking amount input
    try:
        await amount_input.press("Control+a")
    except Exception:
        pass
    await amount_input.press("Backspace")
    await amount_input.type(f"{amount:.2f}")
    await asyncio.sleep(2.0)  # Wait after typing amount
    
    log('# [BTN] Create (tip)')
    btn_create = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Create")')
    await safe_click(btn_create, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(5.0)  # Wait 5 seconds after creating

# ------- Media selection helpers (reuse from your posting bot) -------

async def click_parent_folder(page: Page, parent_folder_name: str) -> bool:
    try:
        log(f"# [BTN] Clicking parent folder: {parent_folder_name}")
        folder_containers = await page.query_selector_all("div.OsMxUq_folder")
        for container in folder_containers:
            try:
                name_elem = await container.query_selector("p.OsMxUq_name")
                if not name_elem: continue
                folder_name = (await name_elem.inner_text()).lower().strip()
                if parent_folder_name.lower() in folder_name:
                    link_elem = await container.query_selector("a.OsMxUq_folderImage")
                    if link_elem:
                        await link_elem.click()
                        log(f"# [SUCCESS] Clicked parent folder: {folder_name}")
                        return True
                    else:
                        await container.click()
                        log(f"# [SUCCESS] Clicked parent folder container: {folder_name}")
                        return True
            except Exception as e:
                log(f"# [DEBUG] Error checking folder container: {e}")
                continue
        log(f"# [ERROR] Parent folder '{parent_folder_name}' not found")
        return False
    except Exception as e:
        log(f"# [ERROR] Error clicking parent folder: {e}")
        return False

async def select_folder(page: Page, cfg: Dict[str, Any], campaign: Dict[str, Any], model: Dict[str, Any]):
    filt = campaign.get("folder_filter", {}) or {}
    name_need = (filt.get("name_contains") or "").lower()
    nested_config = model.get("nested_folders", {})
    parent_folder = nested_config.get("parent") if nested_config else None

    if parent_folder:
        log(f"# [STEP] Nested folder detected - clicking parent '{parent_folder}'")
        await asyncio.sleep(3.0)
        parent_found = await click_parent_folder(page, parent_folder)
        if parent_found:
            await asyncio.sleep(10.0)
            log("# [INFO] Waiting for nested folders to load...")

    log(f"# [STEP] Select folder by name: looking for '{name_need}'")
    await asyncio.sleep(5.0)

    async def try_find():
        available_folders = []
        try:
            folder_containers = await page.query_selector_all("div.OsMxUq_folder")
            for container in folder_containers:
                try:
                    name_elem = await container.query_selector("p.OsMxUq_name")
                    if not name_elem: continue
                    folder_name = (await name_elem.inner_text()).lower().strip()
                    if folder_name: available_folders.append(folder_name)
                    if name_need in folder_name:
                        link_elem = await container.query_selector("a.OsMxUq_folderImage")
                        if link_elem: return link_elem
                        return container
                except Exception:
                    continue
        except Exception as e:
            log(f"# [DEBUG] Error finding folder containers: {e}")

        if available_folders:
            log(f"# [DEBUG] Available media folders: {list(set(available_folders))[:15]}")
        return None

    found = await try_find()
    if not found:
        raise RuntimeError(f"Folder not found. Need name containing: '{name_need}'")
    await asyncio.sleep(1.0)
    await found.click()
    await asyncio.sleep(10.0)
    log(f"# [SUCCESS] Selected folder containing '{name_need}'")

async def scroll_media_grid(page: Page) -> bool:
    print("⏳ Scrolling inside media grid to load more items...")
    grid_selector = 'div.ZjJRiW_gridContainer.abvcKa_grid > div'
    last_count = 0
    attempts_without_new = 0
    max_attempts_without_new = 5
    while attempts_without_new < max_attempts_without_new:
        try:
            await page.wait_for_selector(grid_selector, timeout=30000)
            media_items = await page.query_selector_all(grid_selector)
            current_count = len(media_items)
            if current_count > last_count:
                print(f"📸 Media count increased: {last_count} → {current_count}")
                last_count = current_count
                attempts_without_new = 0
                if media_items:
                    await media_items[-1].scroll_into_view_if_needed()
                    await asyncio.sleep(3)
            else:
                attempts_without_new += 1
                print(f"⚠️ No new media (Try {attempts_without_new}/{max_attempts_without_new})")
                if media_items:
                    await media_items[-1].scroll_into_view_if_needed()
                    await asyncio.sleep(2)
        except Exception as e:
            print(f"⚠️ Error scrolling media grid: {e}")
            break
    print(f"✅ Finished scrolling, found {last_count} media items")
    return last_count > 0

async def select_random_media(page: Page, media_type: str = "photo", count: int = 1) -> bool:
    print(f"🎲 Selecting {count} random {media_type} item(s)...")
    grid_selector = 'div.ZjJRiW_gridContainer.abvcKa_grid > div'
    try:
        await page.wait_for_selector(grid_selector, timeout=45000)
        media_items = await page.query_selector_all(grid_selector)
        if len(media_items) < count:
            print(f"❌ Not enough media items found. Need {count}, found {len(media_items)}")
            return False
        selected_items = random.sample(media_items, count)
        for i, item in enumerate(selected_items):
            await item.click()
            print(f"✅ Selected {media_type} item {i + 1}/{count}")
            await asyncio.sleep(1)
        print(f"✅ Selected {count} random {media_type} item(s)")
        return True
    except Exception as e:
        print(f"❌ Error selecting random media: {e}")
        return False

async def select_random_video(page: Page, media_type: str = "video", count: int = 1) -> bool:
    print(f"🎬 Selecting {count} random {media_type}(s)...")
    grid_selector = 'div.ZjJRiW_gridContainer.abvcKa_grid > div'
    try:
        await page.wait_for_selector(grid_selector, timeout=45000)
        media_items = await page.query_selector_all(grid_selector)
        if len(media_items) < count:
            print(f"❌ Not enough video items found. Need {count}, found {len(media_items)}")
            return False
        selected_items = random.sample(media_items, count)
        for i, item in enumerate(selected_items):
            await item.click()
            print(f"✅ Selected {media_type} item {i + 1}/{count}")
            await asyncio.sleep(1)
        print(f"✅ Selected {count} random {media_type}(s)")
        return True
    except Exception as e:
        print(f"❌ Error selecting random video: {e}")
        return False

async def select_media_step(page: Page, cfg: Dict[str, Any], campaign: Dict[str, Any]):
    n = int(campaign["media"]["count"])
    media_type = (campaign["media"].get("type") or "photo").lower()
    log(f"# [STEP] Load & select {n} media items (type={media_type})")
    await scroll_media_grid(page)
    ok = await (select_random_video if media_type == "video" else select_random_media)(page, media_type, n)
    if not ok:
        raise RuntimeError("Failed to select the requested number of media items")

async def send_media_flow(page: Page, cfg: Dict[str, Any], campaign: Dict[str, Any], model: Dict[str, Any], text: str):
    # Click "Send media" card
    log('# [BTN] Type: Send media')
    send_media = page.locator('div.sqNJyW_card[data-testid="button-card"]:has(.sqNJyW_header:has-text("Send media"))')
    await safe_click(send_media, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(3.0)  # Wait for media form to load
    
    log('# [BTN] Next (after selecting Send media)')
    btn_next = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(btn_next, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(3.0)  # Wait for form to load

    await type_textarea(page, cfg, text)

    # Add media (camera-like empty slot)
    log('# [BTN] Add media (open library)')
    add_media = page.locator('div.xhYF1W_emptySlot.xhYF1W_mediaPreview.xhYF1W_addButton[data-testid="add-media-button"]')
    await safe_click(add_media, retries=cfg["retries"]["clicks"])

    # Select folder + media
    await select_folder(page, cfg, campaign, model)
    await select_media_step(page, cfg, campaign)

    # Next (library)
    log('# [BTN] Next (library after selecting media)')
    next_lib = page.locator('div.abvcKa_desktopButtons >> div._8HzuwW_button._8HzuwW_primary:has-text("Next")')
    await safe_click(next_lib, retries=cfg["retries"]["clicks"])

    # Next (popup to pricing)
    log('# [BTN] Next (to pricing for media DM)')
    next_popup = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Next")')
    await safe_click(next_popup, retries=cfg["retries"]["clicks"])

    # Paid or Free
    paid_cfg = campaign.get("paid", {}) or {}
    if paid_cfg.get("enabled", False):
        price = _clamp_price(paid_cfg.get("price", 5), 5, 2500)
        log('# [FIELD] Paid price input')
        price_input = page.locator('input.KKpREq_input[data-testid="decimal-visual-input"]')
        await price_input.wait_for(state="visible", timeout=cfg["timeouts"]["default_ms"])
        await price_input.click()
        try:
            await price_input.press("Control+a")
        except Exception:
            pass
        await price_input.press("Backspace")
        await price_input.type(f"{price:.2f}")
        
        log('# [BTN] Create (paid media dm)')
        btn_create = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Create")')
        await safe_click(btn_create, retries=cfg["retries"]["clicks"])
        await asyncio.sleep(5.0)  # Wait 5 seconds after creating
    else:
        log('# [BTN] Pricing: Free card')
        free_card = page.locator('div.ztm8ka_expandingButtonCard[data-testid="expanding-button-card"]:has(.ztm8ka_header:has-text("Free"))')
        await safe_click(free_card, retries=cfg["retries"]["clicks"])
        
        log('# [BTN] Create (free media dm)')
        btn_create = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Create")')
        await safe_click(btn_create, retries=cfg["retries"]["clicks"])
        await asyncio.sleep(5.0)  # Wait 5 seconds after creating

# ========================= 8. SCHEDULE + SEND NOW =========================

async def schedule_and_send_now(page: Page, cfg: Dict[str, Any]):
    await asyncio.sleep(5.0)
    
    log('# [BTN] Schedule')
    schedule_btn = page.locator('div.ddvKiq_footer div._8HzuwW_button._8HzuwW_primary._8HzuwW_fullWidth.ddvKiq_button:has-text("Schedule")')
    await safe_click(schedule_btn, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(5.0)  # Wait 5 seconds after clicking Schedule
    
    log('# [BTN] Send now')
    send_now = page.locator('div._8HzuwW_popupButtonContainer button._7lrYkG_button._7lrYkG_primary:has-text("Send now")')
    await safe_click(send_now, retries=cfg["retries"]["clicks"])
    await asyncio.sleep(5.0)  # Wait 5 seconds after clicking Send now
    
    log('# [WAIT] Final wait before browser close')
    await asyncio.sleep(5.0)  # Wait 5 seconds before closing browser

# ========================= 9. ORCHESTRATION =========================

async def run_campaign_for_model(cfg: Dict[str, Any], model: Dict[str, Any], campaign: Dict[str, Any]):
    # Per-campaign wait until time (before we even open a browser)
    await maybe_wait_for_campaign_time(campaign, cfg)

    # Resolve message + archive cfgs
    msg_cfg = resolve_message_cfg(campaign, model, cfg)
    arch_cfg = resolve_message_archive_cfg(campaign, model, cfg)

    # Replenish if source empty, then POP a message
    if msg_cfg.get("excel", {}):
        if replenish_excel_if_empty(msg_cfg["excel"], arch_cfg):
            log("# [STEP] Replenished DM Excel from archive")
        dm_text = pop_message_from_excel(msg_cfg["excel"])
    else:
        # Optional inline fallback (not recommended if you want MOVE semantics)
        inline_vals = (msg_cfg.get("inline") or [])
        dm_text = random.choice([v for v in inline_vals if str(v).strip()]) if inline_vals else ""

    pw, browser, context, page = await open_creators_page(cfg)
    used_text_archived = False
    try:
        await search_and_open_creator(page, cfg, model)
        await open_messages(page, cfg)

        # Audience step
        audience = (campaign.get("audience") or "followers").lower()
        if audience == "followers":
            await audience_followers_only_next(page, cfg)
        elif audience == "fans":
            await audience_fans_only_next(page, cfg)
        elif audience in ("followers_and_fans", "both"):
            await audience_followers_and_fans_next(page, cfg)
        else:
            raise RuntimeError(f"Unknown audience: {audience}")

        # Message type step
        msg_type = (campaign.get("type") or "text").lower()
        if msg_type == "text":
            await text_only_flow(page, cfg, dm_text)
        elif msg_type == "tip":
            amount = _clamp_price((campaign.get("tip") or {}).get("amount", 5), 5, 2500)
            await request_tip_flow(page, cfg, dm_text, amount)
        elif msg_type == "media":
            await send_media_flow(page, cfg, campaign, model, dm_text)
        else:
            raise RuntimeError(f"Unknown campaign type: {msg_type}")

        # Final schedule + send now
        await schedule_and_send_now(page, cfg)

        # Archive used message (only after we've actually created the mass message)
        if dm_text and msg_cfg.get("excel", {}):
            archive_used_message(dm_text, arch_cfg, model["name"], campaign.get("name", "campaign"), msg_cfg["excel"]["path"])
            used_text_archived = True

        # Back to creators (optional)
        await page.goto("https://f2f.com/agency/creators/", wait_until="domcontentloaded")

    except Exception as e:
        fp = await screenshot_on_error(page, prefix=f"dm_{model['name']}_{campaign.get('name','campaign')}")
        log(f"❌ Error: {e}\n{traceback.format_exc()}\nScreenshot: {fp}")
        # if we popped a message but failed before creating, push it back to source? (optional)
        if dm_text and msg_cfg.get("excel", {}) and not used_text_archived:
            try:
                # Put it back at the top
                wb, ws = _ensure_wb_ws(msg_cfg["excel"]["path"], msg_cfg["excel"].get("sheet","Sheet1"))
                col = msg_cfg["excel"].get("column","A")
                current = _read_column_values(msg_cfg["excel"]["path"], msg_cfg["excel"].get("sheet","Sheet1"), col)
                current.insert(0, dm_text)
                _write_column_values(ws, col, current)
                wb.save(msg_cfg["excel"]["path"])
                log("↩️ Restored popped message back to source Excel due to failure.")
            except Exception as ee:
                log(f"⚠️ Failed to restore message: {ee}")
        raise
    finally:
        await close_browser(pw, browser, context)

async def main():
    cfg = load_cfg()
    between_models = cfg.get("pace", {}).get("between_models", {}) or {"mode": "none"}
    between_campaigns_default = cfg.get("pace", {}).get("between_campaigns", {}) or {"mode": "none"}

    models: List[Dict[str, Any]] = cfg.get("models", [])
    for m_i, model in enumerate(models, start=1):
        log("=" * 70)
        log(f"MODEL {m_i}/{len(models)}: {model['name']}")
        log("=" * 70)
        campaigns: List[Dict[str, Any]] = model.get("campaigns", [])

        # allow per-model override for campaign pacing
        between_campaigns = model.get("pace", {}).get("between_campaigns", between_campaigns_default)

        for c_i, campaign in enumerate(campaigns, start=1):
            log(f"--- Campaign {c_i}/{len(campaigns)} :: {campaign.get('name','campaign')} ---")
            await run_campaign_for_model(cfg, model, campaign)

            # pacing between campaigns for this model
            if c_i < len(campaigns):
                sec = _pick_delay_seconds(between_campaigns)
                if sec > 0:
                    log(f"⏳ Waiting {sec}s between campaigns...")
                    await asyncio.sleep(sec)

        # pacing between models
        if m_i < len(models):
            sec = _pick_delay_seconds(between_models)
            if sec > 0:
                log(f"⏳ Waiting {sec}s between models...")
                await asyncio.sleep(sec)

if __name__ == "__main__":
    asyncio.run(main())
